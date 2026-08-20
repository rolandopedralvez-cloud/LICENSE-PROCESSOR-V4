"""app/routers/import_.py — /api/import*, /api/import-flags*. Moved
verbatim from main.py."""
import re
import json
import datetime
from fastapi import APIRouter, HTTPException, Body, Query, Request

from app.config import DB
from app.core import (
    get_conn, cols, clean_payload, role_for, require_permission, log_activity,
    _current_user_info, PROTECTED, IMPORT_CACHE,
)

router = APIRouter(tags=["import"])

def _clean_region(v):
    return "Region II" if v not in (None, "") else v

_CLASS_MAP = {"FB (BWA)": "FB-BWA", "FB(BWA)": "FB-BWA", "FB (WDN)": "FB-WDN"}
def _clean_class(v):
    if v in (None, ""): return v
    return _CLASS_MAP.get(str(v).strip(), str(v).strip())

def _norm_for_compare(v):
    """Make an existing-DB value and an incoming-Excel value comparable:
    None and '' count as the same 'blank', everything else is trimmed text.
    Import files store dates as 'YYYY-MM-DD' strings already, same as the DB,
    so no further date handling is needed here."""
    if v is None:
        return ""
    return str(v).strip()

def _norm_name(v):
    """Loose-normalize a name for similarity comparison: lowercase, collapse
    whitespace, drop common punctuation that varies between typed entries
    ('INC.' vs 'INC', extra commas/periods) without touching real letters."""
    import re
    if not v:
        return ""
    s = str(v).lower().strip()
    s = re.sub(r"[.,]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s

def _best_name_match(name_norm, candidates):
    """candidates: list of (norm_name, record_dict). Returns (record, ratio)
    for the closest match by text similarity, or (None, 0) if nothing is
    close enough to be worth flagging. difflib is stdlib — no new
    dependency needed."""
    import difflib
    best_rec, best_ratio = None, 0.0
    for cand_norm, rec in candidates:
        if not cand_norm:
            continue
        ratio = difflib.SequenceMatcher(None, name_norm, cand_norm).ratio()
        if ratio > best_ratio:
            best_ratio, best_rec = ratio, rec
    return best_rec, best_ratio

def _row_diff(existing, incoming, field_labels):
    """Compare an existing DB record to an incoming imported row, field by
    field, for whichever fields the imported file actually supplied. Returns
    a list of {field, label, existing, incoming} for every field that
    differs — this is the actual 'flag the difference' output."""
    out = []
    for f, v_in in incoming.items():
        if f not in field_labels:
            continue   # only compare real license fields (skip *_raw helpers etc.)
        v_old = existing.get(f)
        if _norm_for_compare(v_old) != _norm_for_compare(v_in):
            out.append({
                "field": f,
                "label": field_labels.get(f, f),
                "existing": v_old if v_old not in (None, "") else None,
                "incoming": v_in if v_in not in (None, "") else None,
            })
    return out

@router.post("/api/import")
def import_preview(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_import")
    """Parse an uploaded Excel (base64) in the export layout. Caches rows and
    flags each one against the existing database: 'new' (no matching license
    number on file), 'duplicate' (matching license number, identical data —
    importing would just create a pointless copy), or 'conflict' (matching
    license number, but the imported row disagrees with what's on file on
    one or more fields — shown with the exact field-by-field difference so
    it can be reviewed before anything is written). Does NOT insert yet."""
    import base64, io, datetime, secrets
    import openpyxl, print_stage

    b64 = data.get("b64", "")
    src_filename = (data.get("filename") or "").strip()
    if not b64:
        raise HTTPException(400, "No file received")
    try:
        raw = base64.b64decode(b64.split(",")[-1])
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read the Excel file: {e}")

    ws = wb["licenses"] if "licenses" in wb.sheetnames else wb.active

    # header (row 1) -> db field, matched by name against the export layout
    header_to_field = {}
    for h, f in zip(print_stage.STAGING_HEADERS, print_stage.STAGING_FIELDS):
        if f is not None:
            header_to_field[str(h).strip().lower()] = f
    col_field = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None: continue
        f = header_to_field.get(str(h).strip().lower())
        if f: col_field[c] = f
    if not col_field:
        raise HTTPException(400, "This file's columns don't match the export layout. "
                                 "Please import a file created by Export.")

    date_fields = print_stage.DATE_FIELDS
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c, f in col_field.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (datetime.datetime, datetime.date)):
                v = v.strftime("%Y-%m-%d")
            rec[f] = v
        # skip blank rows (no license_no and no licensee)
        if not (rec.get("license_no") or rec.get("licensee")):
            continue
        # cleanup, keeping raw
        if "region" in rec:
            rec["region_raw"] = rec.get("region"); rec["region"] = _clean_region(rec.get("region"))
        if "class_of_station" in rec:
            rec["class_of_station_raw"] = rec.get("class_of_station")
            rec["class_of_station"] = _clean_class(rec.get("class_of_station"))
        # the search list's "Type" badge reads license_status — keep it in sync
        # with the Status column so imported records show their type correctly
        if rec.get("status") and not rec.get("license_status"):
            rec["license_status"] = rec["status"]
        rows.append(rec)

    # ---- flag each row against the existing database, by license number ----
    field_labels = {f: h for h, f in zip(print_stage.STAGING_HEADERS, print_stage.STAGING_FIELDS) if f}
    conn = get_conn()
    existing_by_lic = {}
    for lic_no in {r.get("license_no") for r in rows if r.get("license_no")}:
        row = conn.execute(
            "SELECT * FROM licenses WHERE TRIM(license_no) = TRIM(?) ORDER BY id DESC LIMIT 1",
            (lic_no,)).fetchone()
        if row:
            existing_by_lic[str(lic_no).strip()] = dict(row)

    # Name-similarity index for likely-misspelling detection: only used for
    # rows that have NO exact license-number match, to catch the case where
    # the same station was retyped with a slightly different license number
    # or a typo'd name — an exact match on license_no is handled above and
    # doesn't need this slower check.
    NAME_MATCH_THRESHOLD = 0.87
    name_candidates = []
    if any(not (r.get("license_no") or "").strip() or str(r.get("license_no")).strip() not in existing_by_lic for r in rows):
        for row in conn.execute("SELECT id, license_no, licensee, site_name FROM licenses"):
            rd = dict(row)
            name_candidates.append((_norm_name(rd.get("licensee")), rd))
    conn.close()

    flags = []
    new_count = dup_count = conflict_count = misspell_count = 0
    for rec in rows:
        lic_no = str(rec.get("license_no") or "").strip()
        existing = existing_by_lic.get(lic_no) if lic_no else None
        if not existing:
            # no exact license-number match — check if the licensee name is
            # suspiciously close to an existing record (possible misspelling
            # / re-entry under a different license number)
            incoming_norm = _norm_name(rec.get("licensee"))
            match_rec, ratio = (None, 0.0)
            if incoming_norm:
                match_rec, ratio = _best_name_match(incoming_norm, name_candidates)
            if match_rec and ratio >= NAME_MATCH_THRESHOLD and _norm_name(match_rec.get("licensee")) != incoming_norm:
                flags.append({
                    "status": "possible_misspelling", "license_no": rec.get("license_no"),
                    "licensee": rec.get("licensee"), "existing_id": match_rec["id"],
                    "match_confidence": round(ratio * 100),
                    "diff": [
                        {"field": "license_no", "label": "License No.",
                         "existing": match_rec.get("license_no"), "incoming": rec.get("license_no")},
                        {"field": "licensee", "label": "Licensee",
                         "existing": match_rec.get("licensee"), "incoming": rec.get("licensee")},
                    ],
                })
                misspell_count += 1
            else:
                flags.append({"status": "new", "license_no": rec.get("license_no"),
                              "licensee": rec.get("licensee"), "diff": []})
                new_count += 1
            continue
        diff = _row_diff(existing, rec, field_labels)
        if diff:
            flags.append({"status": "conflict", "license_no": rec.get("license_no"),
                          "licensee": rec.get("licensee"), "existing_id": existing["id"], "diff": diff})
            conflict_count += 1
        else:
            flags.append({"status": "duplicate", "license_no": rec.get("license_no"),
                          "licensee": rec.get("licensee"), "existing_id": existing["id"], "diff": []})
            dup_count += 1

    token = secrets.token_urlsafe(12)
    IMPORT_CACHE[token] = {"rows": rows, "flags": flags, "filename": src_filename}
    return {"ok": True, "token": token, "rows": len(rows),
            "summary": {"new": new_count, "duplicate": dup_count, "conflict": conflict_count,
                        "possible_misspelling": misspell_count},
            "flags": flags}

@router.post("/api/import/commit")
def import_commit(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_import")
    """Insert the previously-parsed rows as NEW records (after a backup).
    By default every parsed row is inserted, same as before. Pass
    skip_indices (0-based, matching the 'flags' list from /api/import) to
    leave out specific rows — e.g. the duplicates/conflicts a reviewer chose
    not to bring in."""
    import shutil, os, datetime
    token = data.get("token", "")
    cached = IMPORT_CACHE.get(token)
    if cached is None:
        raise HTTPException(400, "Import session expired — please pick the file again")
    rows = cached["rows"]
    flags = cached.get("flags") or []
    skip_indices = set(data.get("skip_indices") or [])

    # backup the database first
    backup = None
    try:
        if os.path.exists(DB):
            backup = f"telco_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy(DB, backup)
    except Exception:
        backup = None

    import json as _json
    src_filename = cached.get("filename") or "Excel import"
    batch_tag = f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} — {src_filename}"

    conn = get_conn()
    writable = set(cols(conn, "licenses")) - PROTECTED
    added = 0
    skipped = 0
    try:
        for i, rec in enumerate(rows):
            if i in skip_indices:
                skipped += 1
                continue
            payload = {k: v for k, v in rec.items() if k in writable and v is not None}
            if not payload:
                continue
            payload["import_batch"] = batch_tag
            fields = list(payload.keys())
            sql = (f"INSERT INTO licenses ({','.join(fields)}) "
                   f"VALUES ({','.join('?'*len(fields))})")
            conn.execute(sql, [payload[f] for f in fields])
            added += 1
        conn.commit()

        # Persist anything that was actually flagged (duplicate, conflict, or
        # a possible misspelling) so it can still be reviewed later from the
        # Flagged Records view — regardless of whether that particular row
        # was brought in or skipped. Clean "new" rows aren't worth keeping
        # a record of; they're just normal additions.
        info = _current_user_info(request)
        username = info["username"] if info else None
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for i, f in enumerate(flags):
            if f["status"] == "new":
                continue
            action_taken = "skipped" if i in skip_indices else "imported"
            conn.execute(
                "INSERT INTO import_flags (ts, username, status, license_no, licensee, existing_id, diff_json, action_taken) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (now, username, f["status"], f.get("license_no"), f.get("licensee"),
                 f.get("existing_id"), _json.dumps(f.get("diff") or []), action_taken))
        conn.commit()
    finally:
        conn.close()
    IMPORT_CACHE.pop(token, None)
    skip_note = f", skipped {skipped} flagged row(s)" if skipped else ""
    n_flagged = sum(1 for f in flags if f["status"] != "new")
    flagged_note = f" — {n_flagged} flagged for review" if n_flagged else ""
    log_activity(request, "import", detail=f"Imported {added} new record(s) from Excel{skip_note}{flagged_note}" + (f" (backup: {backup})" if backup else ""))
    return {"ok": True, "added": added, "skipped": skipped, "backup": backup}


@router.get("/api/import-flags")
def list_import_flags(status: str = "", resolved: str = "", ignored: str = "no", q: str = "", limit: int = 300, request: Request = None):
    """Browse duplicates/conflicts/possible-misspellings flagged by past
    Imports and the automatic duplicate-license sweep — the persisted record
    behind the Flagged Records view, so this doesn't disappear once the
    one-time review popup is closed. Ignored flags are hidden by default —
    pass ignored=yes or ignored=all to see them."""
    require_permission(request, "can_import")
    import json as _json
    conn = get_conn()
    clauses, params = [], []
    if status:
        clauses.append("status = ?"); params.append(status)
    if resolved == "yes":
        clauses.append("resolved = 1")
    elif resolved == "no":
        clauses.append("resolved = 0")
    if ignored == "yes":
        clauses.append("ignored = 1")
    elif ignored != "all":
        clauses.append("ignored = 0")   # default: hide ignored flags
    if q:
        clauses.append("(license_no LIKE ? OR licensee LIKE ?)")
        params += [f"%{q}%", f"%{q}%"]
    clause = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    rows = conn.execute(
        f"SELECT * FROM import_flags {clause} ORDER BY id DESC LIMIT ?", params + [limit]).fetchall()
    total = conn.execute(f"SELECT COUNT(*) FROM import_flags {clause}", params).fetchone()[0]
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["diff"] = _json.loads(d.pop("diff_json") or "[]")
        except Exception:
            d["diff"] = []
            d.pop("diff_json", None)
        out.append(d)
    return {"count": len(out), "total": total, "results": out}


@router.post("/api/import-flags/{flag_id}/resolve")
def resolve_import_flag(flag_id: int, data: dict = Body(default={}), request: Request = None):
    """Mark a flagged record as reviewed (or un-mark it), e.g. after the
    office has confirmed by hand whether it was really a misspelling, a
    genuine duplicate, or an intentional correction. This does NOT stop the
    automatic sweep from re-flagging it later if the underlying problem is
    still there — use Ignore for that instead."""
    require_permission(request, "can_import")
    import datetime as _dt
    resolved = bool(data.get("resolved", True))
    info = _current_user_info(request)
    username = info["username"] if info else None
    conn = get_conn()
    if not conn.execute("SELECT 1 FROM import_flags WHERE id = ?", (flag_id,)).fetchone():
        conn.close()
        raise HTTPException(404, "Flag not found")
    if resolved:
        conn.execute(
            "UPDATE import_flags SET resolved = 1, resolved_by = ?, resolved_ts = ? WHERE id = ?",
            (username, _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), flag_id))
    else:
        conn.execute(
            "UPDATE import_flags SET resolved = 0, resolved_by = NULL, resolved_ts = NULL WHERE id = ?",
            (flag_id,))
    conn.commit(); conn.close()
    return {"ok": True, "id": flag_id, "resolved": resolved}


@router.post("/api/import-flags/{flag_id}/ignore")
def ignore_import_flag(flag_id: int, data: dict = Body(default={}), request: Request = None):
    """Mark a flag as permanently dismissed — unlike Clear, the underlying
    row stays around so the automatic sweep (see _sweep_duplicate_licenses)
    knows not to raise this exact license number again. Use this for a
    flag that's a false alarm or an intentionally-kept duplicate."""
    require_permission(request, "can_import")
    import datetime as _dt
    ignored = bool(data.get("ignored", True))
    info = _current_user_info(request)
    username = info["username"] if info else None
    conn = get_conn()
    if not conn.execute("SELECT 1 FROM import_flags WHERE id = ?", (flag_id,)).fetchone():
        conn.close()
        raise HTTPException(404, "Flag not found")
    if ignored:
        conn.execute(
            "UPDATE import_flags SET ignored = 1, ignored_by = ?, ignored_ts = ? WHERE id = ?",
            (username, _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), flag_id))
    else:
        conn.execute(
            "UPDATE import_flags SET ignored = 0, ignored_by = NULL, ignored_ts = NULL WHERE id = ?",
            (flag_id,))
    conn.commit(); conn.close()
    return {"ok": True, "id": flag_id, "ignored": ignored}


@router.delete("/api/import-flags/{flag_id}")
def clear_import_flag(flag_id: int, request: Request = None):
    """Permanently remove a flag from the list. Unlike Ignore, if the
    underlying problem (e.g. a duplicate license number) is still present
    in the database, the automatic sweep can raise it again later — Clear
    is for flags that are simply done with, not a statement that the data
    is fine."""
    require_permission(request, "can_import")
    conn = get_conn()
    cur = conn.execute("DELETE FROM import_flags WHERE id = ?", (flag_id,))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "Flag not found")
    return {"ok": True, "id": flag_id, "cleared": True}


def _sweep_duplicate_licenses():
    """Scan the live database for license numbers shared by more than one
    active record — the kind of problem that can slip in outside of Import
    (a typo'd 'Save as New', a manual entry, etc.) and would otherwise go
    unnoticed. Raises a 'database_duplicate' flag for each duplicate group
    that doesn't already have an open (unresolved, non-ignored) flag and
    hasn't been explicitly ignored. Safe to call repeatedly — never creates
    a second flag for the same still-open problem. Returns how many new
    flags it created."""
    import json as _json, datetime as _dt
    conn = get_conn()
    try:
        groups = conn.execute("""
            SELECT TRIM(license_no) AS lic, GROUP_CONCAT(id) AS ids, COUNT(*) AS n,
                   MIN(licensee) AS licensee
            FROM licenses
            WHERE deleted_at IS NULL AND license_no IS NOT NULL AND TRIM(license_no) <> ''
            GROUP BY TRIM(license_no)
            HAVING COUNT(*) > 1
        """).fetchall()
        created = 0
        now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for g in groups:
            lic = g["lic"]
            already = conn.execute(
                "SELECT 1 FROM import_flags WHERE status = 'database_duplicate' AND license_no = ? "
                "AND resolved = 0 AND ignored = 0 LIMIT 1", (lic,)).fetchone()
            if already:
                continue
            ever_ignored = conn.execute(
                "SELECT 1 FROM import_flags WHERE status = 'database_duplicate' AND license_no = ? "
                "AND ignored = 1 LIMIT 1", (lic,)).fetchone()
            if ever_ignored:
                continue
            ids = g["ids"]
            diff = [{"field": "license_no", "label": "License No.",
                     "existing": lic, "incoming": f"Shared by record IDs: {ids}"}]
            conn.execute(
                "INSERT INTO import_flags (ts, username, status, license_no, licensee, existing_id, diff_json, action_taken) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (now, "system (automatic sweep)", "database_duplicate", lic, g["licensee"],
                 None, _json.dumps(diff), "auto_detected"))
            created += 1
        conn.commit()
        return created
    finally:
        conn.close()


@router.post("/api/import-flags/rescan")
def rescan_import_flags(request: Request = None):
    """Run the automatic duplicate-license sweep right now instead of
    waiting for the background timer."""
    require_permission(request, "can_import")
    created = _sweep_duplicate_licenses()
    if created:
        log_activity(request, "import", detail=f"Manual rescan found {created} new duplicate-license flag(s)")
    return {"ok": True, "created": created}


# (SCAN_DIR / SCAN_MODEL used to be defined here — moved to app/core.py,
# see app/routers/scan.py for their users)
