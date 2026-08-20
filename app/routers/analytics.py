"""app/routers/analytics.py — /api/meta, /api/stats, /api/analytics*,
/api/or-batches*, /api/recent, /api/export. Moved verbatim from main.py."""
import io
import datetime
from fastapi import APIRouter, HTTPException, Body, Query, Request
from fastapi.responses import StreamingResponse

from app.config import DB
from app.core import get_conn, role_for, require_permission, log_activity

router = APIRouter(tags=["analytics"])

# NOTE: /api/meta and /api/stats used to live here (moved verbatim from
# main.py) but have been superseded by the SQLAlchemy versions in
# app/routers/meta.py (ported in step 5/6 of MODERNIZATION_PLAN.md — they
# were pure reads with no auth, the lowest-risk place to start). Removed
# here to avoid two competing implementations of the same route; the old
# copy here also had a latent bug (used cols() without importing it).

# ---------------------------------------------------------------- list / search
def _build_analytics_rows(conn):
    """
    One row per unique License No — duplicates/renewal history collapsed to
    the version with the latest Validity To date (that one represents the
    license's current state: its Tech, Channel, Location, and expiry).
    Status is computed by comparing that date to today.
    """
    import datetime
    today = datetime.date.today()

    rows = conn.execute("""
        SELECT id, license_no, tech, config, brgy, town, province,
               validity_to, license_status, status, licensee, class_of_station,
               site_no, site_name, elong_deg, elong_min, elong_sec, nlat_deg, nlat_min, nlat_sec
        FROM licenses
        WHERE deleted_at IS NULL AND license_no IS NOT NULL AND license_no <> ''
    """).fetchall()

    latest = {}   # license_no -> best row so far
    for r in rows:
        lic = r["license_no"]
        cur = latest.get(lic)
        this_date = r["validity_to"] or ""
        if cur is None or this_date > (cur["validity_to"] or ""):
            latest[lic] = r

    out = []
    for lic, r in latest.items():
        vto = r["validity_to"]
        expired = None
        if vto:
            try:
                d = datetime.datetime.strptime(str(vto)[:10], "%Y-%m-%d").date()
                expired = d < today
            except ValueError:
                expired = None
        out.append({
            "id": r["id"],
            "license_no": lic,
            "licensee": r["licensee"],
            "tech": r["tech"],
            "config": r["config"],
            "class_of_station": r["class_of_station"],
            "site_no": r["site_no"],
            "site_name": r["site_name"],
            "brgy": r["brgy"],
            "town": r["town"],
            "province": r["province"],
            "validity_to": vto,
            "status": ("Expired" if expired else "Renewed") if expired is not None else "Unknown",
            "elong_deg": r["elong_deg"], "elong_min": r["elong_min"], "elong_sec": r["elong_sec"],
            "nlat_deg": r["nlat_deg"], "nlat_min": r["nlat_min"], "nlat_sec": r["nlat_sec"],
        })
    out.sort(key=lambda x: (x["province"] or "", x["town"] or "", x["license_no"] or ""))
    return out


PIVOT_DIMS = {
    "tech": "tech",
    "province": "province",
    "town": "town",
    "brgy": "brgy",
}
PIVOT_LABELS = {"tech": "Technology", "province": "Province", "town": "Municipality", "brgy": "Barangay"}

def _pivot_data(conn, x_dim, y_dim, licensee=None):
    if x_dim not in PIVOT_DIMS or y_dim not in PIVOT_DIMS:
        raise HTTPException(400, "Invalid axis — choose from tech, province, town, brgy")
    rows = _build_analytics_rows(conn)
    if licensee:
        rows = [r for r in rows if (r["licensee"] or "") == licensee]
    xk, yk = PIVOT_DIMS[x_dim], PIVOT_DIMS[y_dim]

    x_vals = sorted(set((r[xk] or f"(no {PIVOT_LABELS[x_dim].lower()})") for r in rows))
    grid = {}
    for r in rows:
        yv = r[yk] or f"(no {PIVOT_LABELS[y_dim].lower()})"
        xv = r[xk] or f"(no {PIVOT_LABELS[x_dim].lower()})"
        grid.setdefault(yv, {})
        grid[yv][xv] = grid[yv].get(xv, 0) + 1

    pivot_rows = []
    for yv, counts in sorted(grid.items()):
        pivot_rows.append({"y": yv, "counts": counts, "total": sum(counts.values())})
    col_totals = {xv: 0 for xv in x_vals}
    for r in pivot_rows:
        for xv, n in r["counts"].items():
            if xv in col_totals:
                col_totals[xv] += n
    grand_total = sum(col_totals.values())
    return {
        "x_dim": x_dim, "y_dim": y_dim,
        "x_label": PIVOT_LABELS[x_dim], "y_label": PIVOT_LABELS[y_dim],
        "x_vals": x_vals, "rows": pivot_rows,
        "col_totals": col_totals, "grand_total": grand_total,
    }


@router.get("/api/analytics/pivot")
def get_analytics_pivot(x: str = "tech", y: str = "province", licensee: str = ""):
    """
    Flexible cross-tab: pick any two dimensions (tech, province, town, brgy)
    for the X and Y axes. Cells = count of unique licenses (one-per-license
    rule, same as /api/analytics). Optionally restrict to one carrier.
    """
    conn = get_conn()
    try:
        return _pivot_data(conn, x, y, licensee or None)
    finally:
        conn.close()


@router.post("/api/analytics/pivot/export")
def export_analytics_pivot(x: str = "tech", y: str = "province", licensee: str = "", request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    conn = get_conn()
    try:
        p = _pivot_data(conn, x, y, licensee or None)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pivot"
    if licensee:
        ws.append([f"Carrier: {licensee}"])
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, italic=True)
        ws.append([])
    headers = [p["y_label"]] + p["x_vals"] + ["Total"]
    ws.append(headers)
    header_row = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).font = openpyxl.styles.Font(bold=True)

    for r in p["rows"]:
        row = [r["y"]] + [r["counts"].get(xv, 0) for xv in p["x_vals"]] + [r["total"]]
        ws.append(row)

    total_row = ["TOTAL"] + [p["col_totals"].get(xv, 0) for xv in p["x_vals"]] + [p["grand_total"]]
    ws.append(total_row)
    for c in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=c).font = openpyxl.styles.Font(bold=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    ws.column_dimensions["A"].width = 22
    for i in range(len(p["x_vals"]) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + i)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_Pivot_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _dms_to_dec(deg, mins, secs):
    try:
        return (float(deg or 0) + float(mins or 0)/60 + float(secs or 0)/3600)
    except (TypeError, ValueError):
        return None


def _filter_by_y(rows, y_dim, values):
    yk = PIVOT_DIMS.get(y_dim)
    if not yk:
        raise HTTPException(400, "Invalid axis — choose from tech, province, town, brgy")
    wanted = set(values)
    out = []
    for r in rows:
        yv = r[yk] or f"(no {PIVOT_LABELS[y_dim].lower()})"
        if yv in wanted:
            out.append(r)
    return out


@router.get("/api/analytics/pivot/selected")
def get_pivot_selected(y: str, values: str, licensee: str = ""):
    """Records whose Y-axis value is in the given list (comma-separated) —
    used for 'export selected' and 'pin selected to map' from the pivot table."""
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()
    if licensee:
        rows = [r for r in rows if (r["licensee"] or "") == licensee]
    vals = [v for v in values.split(",") if v != ""]
    filtered = _filter_by_y(rows, y, vals)
    # attach decimal coordinates for easy map pinning
    for r in filtered:
        r["lat"] = _dms_to_dec(r["nlat_deg"], r["nlat_min"], r["nlat_sec"])
        r["lng"] = _dms_to_dec(r["elong_deg"], r["elong_min"], r["elong_sec"])
    return {"count": len(filtered), "rows": filtered}


@router.post("/api/analytics/pivot/selected/export")
def export_pivot_selected(y: str, values: str, licensee: str = "", request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()
    if licensee:
        rows = [r for r in rows if (r["licensee"] or "") == licensee]
    vals = [v for v in values.split(",") if v != ""]
    filtered = _filter_by_y(rows, y, vals)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "selected"
    headers = ["License No","Licensee","Tech","Channel/Config","Class of Station",
               "Site Name","Barangay","Municipality","Province","Validity To","Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for r in filtered:
        ws.append([r["license_no"], r["licensee"], r["tech"], r["config"], r["class_of_station"],
                   r["site_name"], r["brgy"], r["town"], r["province"], r["validity_to"], r["status"]])
    ws.freeze_panes = "A2"
    for i, w in enumerate([18,26,12,14,14,20,16,16,14,12,10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_Selected_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/api/recent")
def get_recent(limit: int = 50):
    """The most recently added/saved records, newest first — for the
    'Recently Added' screen, so nothing gets lost after a busy day of entry."""
    conn = get_conn()
    try:
        rows = conn.execute("""
            SELECT id, license_no, licensee, site_name, brgy, town, province,
                   status, tech, created_at, updated_at, import_batch
            FROM licenses
            WHERE deleted_at IS NULL
            ORDER BY created_at DESC
            LIMIT ?
        """, (limit,)).fetchall()
    finally:
        conn.close()
    return {"records": [dict(r) for r in rows]}


@router.get("/api/or-batches")
def get_or_batches():
    """
    Groups every record by its (OR No., OR Date) pair — since Batch Renew
    writes the exact same OR info to every station in a batch, this pair
    naturally identifies "which stations were renewed together on this
    receipt." Always computed live from the database, so it reflects every
    batch renewal automatically — nothing extra to keep in sync.
    """
    conn = get_conn()
    try:
        rows = conn.execute("""
            SELECT id, license_no, licensee, site_name, brgy, town, province,
                   or_no, or_date, or_amount, rsl_date, status, tech,
                   elong_deg, elong_min, elong_sec, nlat_deg, nlat_min, nlat_sec
            FROM licenses
            WHERE deleted_at IS NULL AND or_no IS NOT NULL AND or_no <> ''
                  AND or_date IS NOT NULL AND or_date <> ''
            ORDER BY or_date DESC, or_no
        """).fetchall()
    finally:
        conn.close()

    groups = {}
    order = []
    for r in rows:
        key = (r["or_no"], r["or_date"])
        if key not in groups:
            groups[key] = {
                "or_no": r["or_no"], "or_date": r["or_date"],
                "or_amount": r["or_amount"], "records": [],
            }
            order.append(key)
        groups[key]["records"].append({
            "id": r["id"], "license_no": r["license_no"], "licensee": r["licensee"],
            "site_name": r["site_name"], "brgy": r["brgy"], "town": r["town"],
            "province": r["province"], "rsl_date": r["rsl_date"], "status": r["status"], "tech": r["tech"],
            "elong_deg": r["elong_deg"], "elong_min": r["elong_min"], "elong_sec": r["elong_sec"],
            "nlat_deg": r["nlat_deg"], "nlat_min": r["nlat_min"], "nlat_sec": r["nlat_sec"],
        })

    batches = [groups[k] for k in order]
    for b in batches:
        b["count"] = len(b["records"])
    return {"total_batches": len(batches), "batches": batches}


@router.post("/api/or-batches/export")
def export_or_batches(request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    data = get_or_batches()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "or_batches"
    headers = ["OR No.", "OR Date", "OR Amount", "# Records", "License No.", "Licensee",
               "Site Name", "Barangay", "Municipality", "Province", "RSL Date", "Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for b in data["batches"]:
        for r in b["records"]:
            ws.append([b["or_no"], b["or_date"], b["or_amount"], b["count"],
                       r["license_no"], r["licensee"], r["site_name"], r["brgy"],
                       r["town"], r["province"], r["rsl_date"], r["status"]])
    ws.freeze_panes = "A2"
    for i, w in enumerate([16,12,14,10,18,24,18,16,16,14,12,10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"OR_Batches_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/api/analytics")
def get_analytics():
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()
    expired_n = sum(1 for r in rows if r["status"] == "Expired")
    renewed_n = sum(1 for r in rows if r["status"] == "Renewed")
    return {"total_unique_licenses": len(rows), "expired": expired_n, "renewed": renewed_n, "rows": rows}


@router.post("/api/analytics/export")
def export_analytics(request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "analytics"
    headers = ["License No","Licensee","Tech","Channel/Config","Class of Station",
               "Barangay","Municipality","Province","Validity To","Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for r in rows:
        ws.append([r["license_no"], r["licensee"], r["tech"], r["config"], r["class_of_station"],
                   r["brgy"], r["town"], r["province"], r["validity_to"], r["status"]])
    ws.freeze_panes = "A2"
    for i, w in enumerate([18,26,12,14,14,16,16,14,12,10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_Analytics_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------- EXPORT to Excel
@router.post("/api/export")
def export_records(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_export")
    """
    Export the given record ids to an .xlsx in the original TELCO_DATABASE layout
    (so it's interchangeable and re-importable). Only reads data; changes nothing.
    """
    import io
    import openpyxl
    from app.legacy import print_stage
    from datetime import datetime

    ids = data.get("ids", [])
    export_all = bool(data.get("all"))

    conn = get_conn()
    if export_all:
        ids = [r[0] for r in conn.execute(
            "SELECT id FROM licenses WHERE deleted_at IS NULL ORDER BY id")]
    if not ids:
        conn.close()
        raise HTTPException(400, "No records to export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "licenses"
    for col, h in enumerate(print_stage.STAGING_HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)

    # which staging columns hold dates, so we can force a date-only display
    # (values are Python datetimes at midnight — without this, Excel shows
    # a 00:00:00 time alongside the date)
    date_cols = {i for i, f in enumerate(print_stage.STAGING_FIELDS, start=1)
                 if f in print_stage.DATE_FIELDS}

    r = 2
    for lic_id in ids:
        row = conn.execute("SELECT * FROM licenses WHERE id = ?", (lic_id,)).fetchone()
        if not row:
            continue
        rec = dict(row)
        for col, value in print_stage.staging_cells(rec):
            cell = ws.cell(row=r, column=col, value=value)
            if col in date_cols:
                cell.number_format = "yyyy-mm-dd"
        r += 1
    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------- PRINT (single record)
