"""app/legacy/services_builtin.py — the "Other Services" feature.

Everything on this app so far (licenses, print, analytics, batch renew...)
is built specifically around Telco RSL licensing. This module adds a
SEPARATE, generic system for every OTHER kind of service the office
handles (water permits, business permits, whatever comes next) --
completely additive, does not touch the `licenses` table or any Telco RSL
route/template. Telco RSL keeps working exactly as it always has.

How it works, end to end:
  1. "Type of Service" and "Nature of Service" are simple named categories
     (service_types / service_natures) -- a small, flat list, picked or
     added on the fly when creating a service.
  2. A "Service" (services table) is one kind of service, e.g. "Water
     Permit". It owns its own custom field layout (`fields`, JSON: label,
     field key, type, and an x/y/w/h position on a canvas -- the SAME
     "design it yourself, drag things around" idea as the Telco print
     Design Mode, just applied to the data-entry form this time).
  3. Each actual application/record for that service is one row in
     service_records, with all its field values in one JSON blob
     (`data`) keyed by field key -- since every service has different
     fields, there's no fixed column list to maintain.
  4. Printing reuses the exact same box/paragraph/run template engine as
     Telco RSL (print_builtin.py) -- see build_pdf() below, which is the
     same idea with field values coming from service_records.data instead
     of a SQL row. Design Mode on the print preview page works exactly
     the same way it does for Telco (drag, resize, undo/redo, colors,
     paper size, grid, layout history, page numbers -- all of it).
  5. Excel export column choice (which fields, which order) and analytics
     widgets (count-by, sum, trend) are both just small JSON configs
     stored on the service row, editable from the service's settings.
"""
import os
import io
import json
import base64
import datetime
import sqlite3
import xml.sax.saxutils as saxutils
import re

from app.config import DB

_HEX_RE = re.compile(r"^#[0-9a-fA-F]{6}$")
FIELD_TYPES = ("text", "textarea", "number", "date", "select", "checkbox", "location")

# Where uploaded record attachments (photos, scanned IDs, supporting docs)
# are stored -- same "real files on disk, path in the DB" pattern as
# app/routers/scan.py's SCAN_DIR, just namespaced for this feature.
ATTACHMENT_DIR_NAME = "service_attachments"


def _project_root():
    # this file lives in app/legacy/ -- attachments live two levels up,
    # next to telco.db/start.bat (same convention as print_builtin.py's
    # _project_root()).
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


def get_conn():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def _cols(conn, table):
    return [r["name"] for r in conn.execute(f"PRAGMA table_info({table})")]


def ensure_services_schema():
    conn = get_conn()
    conn.execute("""CREATE TABLE IF NOT EXISTS service_types (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS service_natures (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        type_id INTEGER,
        nature_id INTEGER,
        fields TEXT NOT NULL DEFAULT '[]',
        excel_columns TEXT NOT NULL DEFAULT '[]',
        print_template TEXT NOT NULL DEFAULT '{"page":{"width_pt":612,"height_pt":792},"image":null,"boxes":[]}',
        print_template_default TEXT NOT NULL DEFAULT '{"page":{"width_pt":612,"height_pt":792},"image":null,"boxes":[]}',
        analytics_config TEXT NOT NULL DEFAULT '[]',
        created_at TEXT,
        archived INTEGER NOT NULL DEFAULT 0
    )""")
    # Columns added after the first release of this feature -- added the
    # same safe way every other table in this app is upgraded (check, then
    # ALTER TABLE ADD COLUMN if missing), so upgrading never loses an
    # existing services table.
    svc_cols = _cols(conn, "services")
    if "board_field" not in svc_cols:
        conn.execute("ALTER TABLE services ADD COLUMN board_field TEXT;")
    if "expiry_field" not in svc_cols:
        conn.execute("ALTER TABLE services ADD COLUMN expiry_field TEXT;")
    if "restricted_users" not in svc_cols:
        conn.execute("ALTER TABLE services ADD COLUMN restricted_users TEXT NOT NULL DEFAULT '[]';")

    conn.execute("""CREATE TABLE IF NOT EXISTS service_records (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_id INTEGER NOT NULL,
        data TEXT NOT NULL DEFAULT '{}',
        created_at TEXT,
        updated_at TEXT,
        deleted_at TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS service_print_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        service_id INTEGER NOT NULL,
        name TEXT,
        saved_at TEXT,
        data TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS service_attachments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        record_id INTEGER NOT NULL,
        filename TEXT,
        stored_path TEXT,
        mime_type TEXT,
        uploaded_by TEXT,
        uploaded_at TEXT
    )""")
    conn.commit()
    conn.close()


def _now():
    return datetime.datetime.now().isoformat(timespec="seconds")


# ---------------------------------------------------------------- categories
def list_types():
    conn = get_conn()
    rows = [dict(r) for r in conn.execute("SELECT * FROM service_types ORDER BY name")]
    conn.close()
    return rows


def add_type(name):
    name = (name or "").strip()
    if not name:
        raise ValueError("Type name is required")
    conn = get_conn()
    try:
        conn.execute("INSERT OR IGNORE INTO service_types (name) VALUES (?)", (name,))
        conn.commit()
        row = conn.execute("SELECT * FROM service_types WHERE name = ?", (name,)).fetchone()
        return dict(row)
    finally:
        conn.close()


def list_natures():
    conn = get_conn()
    rows = [dict(r) for r in conn.execute("SELECT * FROM service_natures ORDER BY name")]
    conn.close()
    return rows


def add_nature(name):
    name = (name or "").strip()
    if not name:
        raise ValueError("Nature name is required")
    conn = get_conn()
    try:
        conn.execute("INSERT OR IGNORE INTO service_natures (name) VALUES (?)", (name,))
        conn.commit()
        row = conn.execute("SELECT * FROM service_natures WHERE name = ?", (name,)).fetchone()
        return dict(row)
    finally:
        conn.close()


# ---------------------------------------------------------------- services
def _service_out(row):
    d = dict(row)
    for k in ("fields", "excel_columns", "print_template", "print_template_default", "analytics_config", "restricted_users"):
        try:
            d[k] = json.loads(d[k]) if d[k] else ([] if k != "print_template" and k != "print_template_default" else {})
        except Exception:
            d[k] = [] if k not in ("print_template", "print_template_default") else {"page": {"width_pt": 612, "height_pt": 792}, "image": None, "boxes": []}
    return d


# ---------------------------------------------------------------- access control
def user_can_access_service(service, username, role):
    """Per-service access restriction ("Restrict Access" on the Other
    Services list, super-admin only to set). Empty restricted_users list =
    everyone with the base can_create/can_edit/etc permission can use it
    (the default, matches every service created before this existed).
    super_admin can always see everything, same as every other admin
    override in this app."""
    if role == "super_admin":
        return True
    restricted = service.get("restricted_users") or []
    if not restricted:
        return True
    return username in restricted


def save_restricted_users(service_id, usernames):
    conn = get_conn()
    conn.execute("UPDATE services SET restricted_users = ? WHERE id = ?",
                 (json.dumps(usernames or []), service_id))
    conn.commit()
    conn.close()


def list_services(include_archived=False, username=None, role=None):
    conn = get_conn()
    q = "SELECT * FROM services"
    if not include_archived:
        q += " WHERE archived = 0"
    q += " ORDER BY name"
    rows = [_service_out(r) for r in conn.execute(q)]
    conn.close()
    if username is not None or role is not None:
        rows = [r for r in rows if user_can_access_service(r, username, role)]
    return rows


def get_service(service_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM services WHERE id = ?", (service_id,)).fetchone()
    conn.close()
    return _service_out(row) if row else None


def create_service(name, type_id=None, nature_id=None):
    name = (name or "").strip()
    if not name:
        raise ValueError("Service name is required")
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO services (name, type_id, nature_id, created_at) VALUES (?, ?, ?, ?)",
        (name, type_id, nature_id, _now()),
    )
    conn.commit()
    sid = cur.lastrowid
    conn.close()
    return get_service(sid)


def update_service_meta(service_id, name=None, type_id=None, nature_id=None, archived=None,
                         board_field=None, expiry_field=None):
    conn = get_conn()
    sets, vals = [], []
    if name is not None:
        sets.append("name = ?"); vals.append(name)
    if type_id is not None:
        sets.append("type_id = ?"); vals.append(type_id)
    if nature_id is not None:
        sets.append("nature_id = ?"); vals.append(nature_id)
    if archived is not None:
        sets.append("archived = ?"); vals.append(1 if archived else 0)
    if board_field is not None:
        sets.append("board_field = ?"); vals.append(board_field or None)
    if expiry_field is not None:
        sets.append("expiry_field = ?"); vals.append(expiry_field or None)
    if sets:
        vals.append(service_id)
        conn.execute(f"UPDATE services SET {', '.join(sets)} WHERE id = ?", vals)
        conn.commit()
    conn.close()
    return get_service(service_id)


def _valid_fields(fields):
    """Sanitizes an incoming field-layout list -- generates missing keys
    from the label, drops anything malformed, keeps only known field
    types. Same defensive posture as the Telco template save (never trust
    the browser's JSON wholesale)."""
    out = []
    seen_keys = set()
    for i, f in enumerate(fields or []):
        label = str(f.get("label") or f"Field {i+1}").strip()
        key = str(f.get("key") or "").strip()
        if not key:
            key = re.sub(r"[^a-z0-9_]", "_", label.lower()).strip("_") or f"field_{i+1}"
        base_key = key
        n = 2
        while key in seen_keys:
            key = f"{base_key}_{n}"; n += 1
        seen_keys.add(key)
        ftype = f.get("type") if f.get("type") in FIELD_TYPES else "text"
        out.append({
            "key": key, "label": label, "type": ftype,
            "x": float(f.get("x", 20)), "y": float(f.get("y", 20)),
            "w": float(f.get("w", 220)), "h": float(f.get("h", 40)),
            "required": bool(f.get("required")),
            "options": str(f.get("options") or ""),
        })
    return out


def save_fields(service_id, fields):
    fields = _valid_fields(fields)
    conn = get_conn()
    conn.execute("UPDATE services SET fields = ? WHERE id = ?", (json.dumps(fields), service_id))
    conn.commit()
    conn.close()
    return fields


def save_excel_columns(service_id, columns):
    conn = get_conn()
    conn.execute("UPDATE services SET excel_columns = ? WHERE id = ?", (json.dumps(columns or []), service_id))
    conn.commit()
    conn.close()


def save_analytics_config(service_id, config):
    conn = get_conn()
    conn.execute("UPDATE services SET analytics_config = ? WHERE id = ?", (json.dumps(config or []), service_id))
    conn.commit()
    conn.close()


# ---------------------------------------------------------------- print template
def save_print_template(service_id, data):
    conn = get_conn()
    conn.execute("UPDATE services SET print_template = ? WHERE id = ?", (json.dumps(data), service_id))
    conn.commit()
    conn.close()


def save_print_template_as_default(service_id):
    svc = get_service(service_id)
    conn = get_conn()
    conn.execute("UPDATE services SET print_template_default = ? WHERE id = ?",
                 (json.dumps(svc["print_template"]), service_id))
    conn.commit()
    conn.close()


def load_print_history(service_id):
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        "SELECT id, name, saved_at FROM service_print_history WHERE service_id = ? ORDER BY id DESC", (service_id,))]
    conn.close()
    return rows


def add_print_history_entry(service_id, name, data):
    conn = get_conn()
    conn.execute(
        "INSERT INTO service_print_history (service_id, name, saved_at, data) VALUES (?, ?, ?, ?)",
        (service_id, (name or "Untitled layout").strip()[:80], _now(), json.dumps(data)),
    )
    # keep newest 30 per service
    conn.execute("""DELETE FROM service_print_history WHERE service_id = ? AND id NOT IN (
        SELECT id FROM service_print_history WHERE service_id = ? ORDER BY id DESC LIMIT 30)""",
        (service_id, service_id))
    conn.commit()
    row = conn.execute("SELECT id, name, saved_at FROM service_print_history WHERE service_id = ? ORDER BY id DESC LIMIT 1", (service_id,)).fetchone()
    conn.close()
    return dict(row)


def get_print_history_entry(service_id, entry_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM service_print_history WHERE service_id = ? AND id = ?", (service_id, entry_id)).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    d["data"] = json.loads(d["data"])
    return d


def delete_print_history_entry(service_id, entry_id):
    conn = get_conn()
    cur = conn.execute("DELETE FROM service_print_history WHERE service_id = ? AND id = ?", (service_id, entry_id))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok


# ---------------------------------------------------------------- records
def format_value(field_def, value):
    if value in (None, ""):
        return ""
    if field_def and field_def.get("type") == "date":
        s = str(value)
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            try:
                d = datetime.date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
                return d.strftime("%B %-d, %Y") if os.name != "nt" else d.strftime("%B %d, %Y").replace(" 0", " ")
            except ValueError:
                return s
        return s
    return str(value)


def formatted_fields(service, record):
    fmap = {f["key"]: f for f in service["fields"]}
    data = record.get("data") or {}
    return {k: format_value(fmap.get(k), v) for k, v in data.items()}


def list_records(service_id, q=""):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM service_records WHERE service_id = ? AND deleted_at IS NULL ORDER BY id DESC",
        (service_id,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["data"] = json.loads(d["data"]) if d["data"] else {}
        except Exception:
            d["data"] = {}
        if q:
            hay = " ".join(str(v) for v in d["data"].values()).lower()
            if q.lower() not in hay:
                continue
        out.append(d)
    return out


def get_record(service_id, record_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM service_records WHERE service_id = ? AND id = ? AND deleted_at IS NULL",
                        (service_id, record_id)).fetchone()
    conn.close()
    if not row:
        return None
    d = dict(row)
    try:
        d["data"] = json.loads(d["data"]) if d["data"] else {}
    except Exception:
        d["data"] = {}
    return d


def create_record(service_id, data):
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO service_records (service_id, data, created_at, updated_at) VALUES (?, ?, ?, ?)",
        (service_id, json.dumps(data or {}), _now(), _now()))
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    return get_record(service_id, rid)


def update_record(service_id, record_id, data):
    conn = get_conn()
    conn.execute("UPDATE service_records SET data = ?, updated_at = ? WHERE service_id = ? AND id = ?",
                 (json.dumps(data or {}), _now(), service_id, record_id))
    conn.commit()
    conn.close()
    return get_record(service_id, record_id)


def delete_record(service_id, record_id):
    conn = get_conn()
    cur = conn.execute("UPDATE service_records SET deleted_at = ? WHERE service_id = ? AND id = ? AND deleted_at IS NULL",
                        (_now(), service_id, record_id))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok


# ---------------------------------------------------------------- trash
def list_trash(service_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM service_records WHERE service_id = ? AND deleted_at IS NOT NULL ORDER BY deleted_at DESC",
        (service_id,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d["data"] = json.loads(d["data"]) if d["data"] else {}
        except Exception:
            d["data"] = {}
        out.append(d)
    return out


def restore_record(service_id, record_id):
    conn = get_conn()
    cur = conn.execute(
        "UPDATE service_records SET deleted_at = NULL WHERE service_id = ? AND id = ? AND deleted_at IS NOT NULL",
        (service_id, record_id))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return ok


def purge_record(service_id, record_id):
    """Permanently delete one trashed record -- also removes its
    attachments (files on disk and their DB rows), same as licenses'
    purge does for the real database."""
    conn = get_conn()
    cur = conn.execute(
        "SELECT stored_path FROM service_attachments WHERE record_id = ?", (record_id,))
    paths = [r["stored_path"] for r in cur.fetchall()]
    conn.execute("DELETE FROM service_attachments WHERE record_id = ?", (record_id,))
    cur = conn.execute(
        "DELETE FROM service_records WHERE service_id = ? AND id = ? AND deleted_at IS NOT NULL",
        (service_id, record_id))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    if ok:
        for p in paths:
            try:
                if p and os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass
    return ok


def empty_trash(service_id):
    conn = get_conn()
    rows = conn.execute(
        "SELECT id FROM service_records WHERE service_id = ? AND deleted_at IS NOT NULL", (service_id,)).fetchall()
    conn.close()
    n = 0
    for r in rows:
        if purge_record(service_id, r["id"]):
            n += 1
    return n


# ---------------------------------------------------------------- attachments
def _attachment_dir():
    d = os.path.join(_project_root(), ATTACHMENT_DIR_NAME)
    os.makedirs(d, exist_ok=True)
    return d


def add_attachment(record_id, b64_data_uri, filename, username):
    """Same base64-in, real-file-on-disk-out pattern as
    app/routers/scan.py's upload_scan -- keeps the database itself small
    and lets a person just open the file normally on the PC if they ever
    need to."""
    import re as _re
    header, _, payload = b64_data_uri.partition(",")
    mime_type = "application/octet-stream"
    m = _re.match(r"data:([\w/\-.+]+);base64", header)
    if m:
        mime_type = m.group(1)
    elif not payload:
        payload = b64_data_uri
    try:
        file_bytes = base64.b64decode(payload or b64_data_uri)
    except Exception:
        raise ValueError("Could not read that file")

    safe_name = _re.sub(r"[^\w.\-]", "_", (filename or "attachment").strip())[:120] or "attachment"
    stored_name = f"{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{record_id}_{safe_name}"
    stored_path = os.path.join(_attachment_dir(), stored_name)
    with open(stored_path, "wb") as f:
        f.write(file_bytes)

    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO service_attachments (record_id, filename, stored_path, mime_type, uploaded_by, uploaded_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (record_id, filename or safe_name, stored_path, mime_type, username, _now()))
    conn.commit()
    aid = cur.lastrowid
    conn.close()
    return {"id": aid, "filename": filename or safe_name, "mime_type": mime_type, "uploaded_at": _now()}


def list_attachments(record_id):
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        "SELECT id, filename, mime_type, uploaded_by, uploaded_at FROM service_attachments "
        "WHERE record_id = ? ORDER BY id DESC", (record_id,))]
    conn.close()
    return rows


def get_attachment(attachment_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM service_attachments WHERE id = ?", (attachment_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_attachment(attachment_id):
    att = get_attachment(attachment_id)
    if not att:
        return False
    conn = get_conn()
    conn.execute("DELETE FROM service_attachments WHERE id = ?", (attachment_id,))
    conn.commit()
    conn.close()
    try:
        if att["stored_path"] and os.path.exists(att["stored_path"]):
            os.remove(att["stored_path"])
    except Exception:
        pass
    return True


# ---------------------------------------------------------------- board (status columns)
def compute_board(service):
    """Groups every (non-deleted) record by the service's chosen
    board_field -- e.g. a Status dropdown -- into columns, the same idea
    as a simple Trello-style board. Records with nothing in that field
    land in a "(blank)" column rather than disappearing."""
    field_key = service.get("board_field")
    if not field_key:
        return {"field": None, "columns": []}
    fmap = {f["key"]: f for f in service["fields"]}
    label_field = service["fields"][0]["key"] if service["fields"] else None
    records = list_records(service["id"])
    columns = {}
    order = []
    # Pre-seed columns from the field's own dropdown choices (if it's a
    # select field) so empty statuses still show up as a column.
    fdef = fmap.get(field_key)
    if fdef and fdef.get("type") == "select":
        for opt in (fdef.get("options") or "").split("\n"):
            opt = opt.strip()
            if opt and opt not in columns:
                columns[opt] = []; order.append(opt)
    for r in records:
        val = str(r["data"].get(field_key) or "(blank)")
        if val not in columns:
            columns[val] = []; order.append(val)
        columns[val].append({
            "id": r["id"],
            "title": r["data"].get(label_field, "") if label_field else "",
        })
    return {"field": field_key, "columns": [{"name": k, "cards": columns[k]} for k in order]}


# ---------------------------------------------------------------- expiry tracking
def compute_expiring(service, within_days=30):
    """Records whose expiry_field date falls within the next `within_days`
    days (or has already passed), for the "Expiring Soon" panel on the
    Records page. Mirrors the plain-string YYYY-MM-DD date handling the
    rest of this module already uses (format_value/_format_date)."""
    field_key = service.get("expiry_field")
    if not field_key:
        return {"field": None, "expiring": [], "expired": []}
    label_field = service["fields"][0]["key"] if service["fields"] else None
    today = datetime.date.today()
    horizon = today + datetime.timedelta(days=within_days)
    expiring, expired = [], []
    for r in list_records(service["id"]):
        raw = r["data"].get(field_key)
        if not raw or len(str(raw)) < 10:
            continue
        s = str(raw)
        try:
            d = datetime.date(int(s[0:4]), int(s[5:7]), int(s[8:10]))
        except (ValueError, IndexError):
            continue
        item = {"id": r["id"], "title": r["data"].get(label_field, "") if label_field else "", "date": s}
        if d < today:
            expired.append(item)
        elif d <= horizon:
            expiring.append(item)
    expiring.sort(key=lambda x: x["date"])
    expired.sort(key=lambda x: x["date"])
    return {"field": field_key, "expiring": expiring, "expired": expired}


# ---------------------------------------------------------------- bulk import
def import_records_from_excel(service, file_bytes):
    """Bulk-loads records from an .xlsx: the first row is treated as
    headers and matched (case-insensitively, ignoring extra spaces)
    against each field's label. Unmatched columns are listed back but
    otherwise ignored -- nothing guesses at a mapping it isn't sure of."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"created": 0, "skipped": 0, "unmatched_columns": []}
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    label_to_key = {f["label"].strip().lower(): f["key"] for f in service["fields"]}
    col_key = [label_to_key.get(h.lower()) for h in header]
    unmatched = [h for h, k in zip(header, col_key) if h and not k]

    created, skipped = 0, 0
    for row in rows[1:]:
        if row is None or all(v in (None, "") for v in row):
            continue
        data = {}
        for i, val in enumerate(row):
            if i >= len(col_key) or not col_key[i] or val in (None, ""):
                continue
            data[col_key[i]] = str(val)
        if not data:
            skipped += 1
            continue
        create_record(service["id"], data)
        created += 1
    return {"created": created, "skipped": skipped, "unmatched_columns": unmatched}


# ---------------------------------------------------------------- Excel export
def build_excel(service):
    import openpyxl
    fmap = {f["key"]: f for f in service["fields"]}
    columns = service.get("excel_columns") or [f["key"] for f in service["fields"]]
    columns = [c for c in columns if c in fmap] or [f["key"] for f in service["fields"]]
    records = list_records(service["id"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = service["name"][:31] or "Records"
    for i, key in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=i, value=fmap[key]["label"])
        cell.font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 22
    for r, rec in enumerate(records, start=2):
        for i, key in enumerate(columns, start=1):
            ws.cell(row=r, column=i, value=format_value(fmap.get(key), rec["data"].get(key)))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- map
def compute_map_points(service):
    """Every record that has at least one "location"-type field filled in
    (stored as "lat,lng" text -- see the field designer's Location field
    type), for the per-service Map page. A service can have more than one
    location field (e.g. two site addresses); each filled one becomes its
    own pin."""
    loc_fields = [f for f in service["fields"] if f["type"] == "location"]
    if not loc_fields:
        return {"fields": [], "points": []}
    label_field = service["fields"][0]["key"] if service["fields"] else None
    points = []
    for r in list_records(service["id"]):
        title = r["data"].get(label_field, "") if label_field else ""
        for f in loc_fields:
            raw = (r["data"].get(f["key"]) or "").strip()
            if not raw or "," not in raw:
                continue
            try:
                lat_s, lng_s = raw.split(",", 1)
                lat, lng = float(lat_s.strip()), float(lng_s.strip())
            except ValueError:
                continue
            points.append({"record_id": r["id"], "field": f["key"], "field_label": f["label"],
                            "title": title, "lat": lat, "lng": lng})
    return {"fields": [f["key"] for f in loc_fields], "points": points}


# ---------------------------------------------------------------- analytics
def compute_analytics(service):
    """Runs the service's saved widget configs against its records. Each
    widget is one of:
      count_by  -- bar list of how many records have each value of `field`
      sum       -- total of a number field
      trend     -- record count per month, by a date field
    Kept deliberately simple (no charting library) so this can never be
    blocked by a CDN the sandbox/office network doesn't allow -- the same
    reasoning the rest of this app's analytics already follows."""
    records = list_records(service["id"])
    out = []
    for widget in service.get("analytics_config") or []:
        wtype = widget.get("type")
        field = widget.get("field")
        label = widget.get("label") or field
        if wtype == "count_by":
            counts = {}
            for r in records:
                v = str(r["data"].get(field) or "(blank)")
                counts[v] = counts.get(v, 0) + 1
            items = sorted(counts.items(), key=lambda kv: -kv[1])[:20]
            maxv = max([c for _, c in items], default=1)
            out.append({"type": "count_by", "label": label, "field": field,
                        "items": [{"name": n, "count": c, "pct": round(100 * c / maxv)} for n, c in items]})
        elif wtype == "sum":
            total = 0.0
            for r in records:
                try:
                    total += float(r["data"].get(field) or 0)
                except (TypeError, ValueError):
                    pass
            out.append({"type": "sum", "label": label, "field": field, "total": total, "count": len(records)})
        elif wtype == "trend":
            months = {}
            for r in records:
                v = str(r["data"].get(field) or "")
                if len(v) >= 7 and v[4] == "-":
                    months[v[:7]] = months.get(v[:7], 0) + 1
            items = sorted(months.items())[-12:]
            maxv = max([c for _, c in items], default=1)
            out.append({"type": "trend", "label": label, "field": field,
                        "items": [{"name": n, "count": c, "pct": round(100 * c / maxv)} for n, c in items]})
    return {"widgets": out, "total_records": len(records)}


# ---------------------------------------------------------------- PDF export
def build_pdf(service, record, batch_index=1, batch_total=1):
    """Same rendering engine as app/legacy/print_builtin.py's build_pdf --
    box/paragraph/run positions from the template, field values resolved
    against this record's data instead of a SQL row. Kept as a parallel
    copy (not a shared import) on purpose, matching the "keep Telco
    separate" decision -- so nothing here can ever affect the Telco RSL
    print path, and vice versa."""
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    values = formatted_fields(service, record)
    tpl = service["print_template"]
    page_w = tpl.get("page", {}).get("width_pt", 612)
    page_h = tpl.get("page", {}).get("height_pt", 792)

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(page_w, page_h))

    img_cfg = tpl.get("image")
    if img_cfg and img_cfg.get("data"):
        try:
            header, b64data = img_cfg["data"].split(",", 1)
            img_bytes = base64.b64decode(b64data)
            img = ImageReader(io.BytesIO(img_bytes))
            c.drawImage(img, img_cfg["x"], page_h - img_cfg["y"] - img_cfg["h"],
                        width=img_cfg["w"], height=img_cfg["h"], mask="auto")
        except Exception:
            pass

    ALIGN = {"left": TA_LEFT, "center": TA_CENTER, "right": TA_RIGHT}

    for box in tpl.get("boxes", []):
        cursor_y_top = page_h - box["y"]
        leading_mult = box.get("leading") or 1.2
        for para in box.get("paragraphs", []):
            markup_parts = []
            base_size = 10
            base_font = "Helvetica"
            for run in para.get("runs", []):
                if run.get("type") == "br":
                    markup_parts.append("<br/>")
                    continue
                if run.get("type") == "field":
                    text = values.get(run.get("field"), "")
                elif run.get("type") == "pagenum":
                    text = f"({batch_index}/{batch_total})"
                else:
                    text = run.get("text", "")
                if not text:
                    continue
                base_size = run.get("size", base_size)
                base_font = run.get("font", "Helvetica") or "Helvetica"
                if base_font not in ("Helvetica", "Times-Roman", "Courier"):
                    base_font = "Helvetica"
                escaped = saxutils.escape(text).replace("\n", "<br/>")
                if run.get("bold"):
                    escaped = f"<b>{escaped}</b>"
                color = run.get("color")
                color_attr = f' color="{color}"' if color and _HEX_RE.match(color) else ""
                markup_parts.append(f'<font face="{base_font}" size="{run.get("size", base_size)}"{color_attr}>{escaped}</font>')
            markup = "".join(markup_parts)
            if not markup.strip():
                continue
            style = ParagraphStyle("box", fontName=base_font, fontSize=base_size,
                                    leading=base_size * leading_mult,
                                    alignment=ALIGN.get(para.get("align", "left"), TA_LEFT))
            p = Paragraph(markup, style)
            w, h = p.wrap(box["w"], box["h"])
            cursor_y_top -= h
            p.drawOn(c, box["x"], cursor_y_top)

    c.showPage()
    c.save()
    return buf.getvalue()
