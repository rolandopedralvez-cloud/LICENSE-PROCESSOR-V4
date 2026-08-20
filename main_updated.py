"""
main.py  —  NTC Region II Telco Database backend (FastAPI)

Run:  uvicorn main:app --reload
Then open:  http://127.0.0.1:8000/docs   (interactive test page)

This backend serves the DATA endpoints (list / search / load / create /
update / delete + payments + stats). The print & preview endpoints are
added in Phase 5, once the printing template is mapped.
"""

import os
import sqlite3
import hashlib
import secrets
from datetime import datetime
from fastapi import FastAPI, HTTPException, Body, Query, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

DB = "telco.db"

app = FastAPI(title="NTC R02 Telco Database", version="1.0")

# Allow the HTML GUI (opened as a local file or another port) to call this API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_methods=["*"], allow_headers=["*"],
)

# ---------------------------------------------------------------- helpers
def get_conn():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

def cols(conn, table):
    return [r["name"] for r in conn.execute(f"PRAGMA table_info({table})")]

# columns the client is never allowed to set directly
PROTECTED = {"id", "created_at", "updated_at"}

def clean_payload(conn, table, data: dict):
    """Keep only real, writable columns from the incoming JSON."""
    valid = set(cols(conn, table)) - PROTECTED
    return {k: v for k, v in data.items() if k in valid}

# ---------------------------------------------------------------- login / auth
PBKDF_ROUNDS = 200_000
TOKENS = {}   # token -> {"username":..., "role":..., "created":..., "last_seen":...}
              # (in memory; cleared on restart)

# Sessions expire two ways: if nothing happens for TOKEN_IDLE_TIMEOUT_SECONDS
# (person walked away / left it signed in on a shared PC), or once
# TOKEN_ABSOLUTE_TIMEOUT_SECONDS is reached regardless of activity (so a
# token that leaked somehow doesn't stay valid forever). Both are generous
# for an office tool -- this isn't meant to log anyone out mid-task.
TOKEN_IDLE_TIMEOUT_SECONDS = 4 * 60 * 60       # 4 hours of no activity
TOKEN_ABSOLUTE_TIMEOUT_SECONDS = 16 * 60 * 60  # 16 hours since sign-in, max

def hash_pw(password, salt):
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, PBKDF_ROUNDS)

def _new_token_entry(username, role):
    import time
    now = time.time()
    return {"username": username, "role": role, "created": now, "last_seen": now}

def _get_token_info(token):
    """Look up a token, evicting it if it's expired. Returns None if the
    token is missing or expired. On success, refreshes last_seen so the
    idle-timeout clock resets on every authenticated request."""
    import time
    info = TOKENS.get(token)
    if not info:
        return None
    now = time.time()
    if now - info.get("created", now) > TOKEN_ABSOLUTE_TIMEOUT_SECONDS:
        TOKENS.pop(token, None)
        return None
    if now - info.get("last_seen", now) > TOKEN_IDLE_TIMEOUT_SECONDS:
        TOKENS.pop(token, None)
        return None
    info["last_seen"] = now
    return info

def role_for(request):
    token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
    info = _get_token_info(token)
    return info["role"] if info else None

def any_user_exists():
    conn = get_conn()
    n = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()
    return n > 0

def ensure_schema():
    """Create/upgrade tables that may be missing."""
    conn = get_conn()
    if "renewed_from" not in cols(conn, "licenses"):
        conn.execute("ALTER TABLE licenses ADD COLUMN renewed_from INTEGER;")
    if "deleted_at" not in cols(conn, "licenses"):
        conn.execute("ALTER TABLE licenses ADD COLUMN deleted_at TEXT;")
    conn.execute("""CREATE TABLE IF NOT EXISTS users (
        username TEXT PRIMARY KEY,
        salt     BLOB NOT NULL,
        pwhash   BLOB NOT NULL
    );""")
    # add role column if this is an older users table (existing user = master/admin)
    if "role" not in cols(conn, "users"):
        conn.execute("ALTER TABLE users ADD COLUMN role TEXT DEFAULT 'admin';")
    # older versions used the role name 'admin' for the top role; normalize it
    # to 'super_admin' so accounts created before this update still get full access
    conn.execute("UPDATE users SET role = 'super_admin' WHERE role = 'admin';")
    # optional 4-digit PIN for quick sign-in, alongside the regular password
    if "pin_salt" not in cols(conn, "users"):
        conn.execute("ALTER TABLE users ADD COLUMN pin_salt BLOB;")
    if "pin_hash" not in cols(conn, "users"):
        conn.execute("ALTER TABLE users ADD COLUMN pin_hash BLOB;")
    # per-user feature checklist for "user"-role accounts (super_admin always
    # has everything regardless of this). Stored as a JSON array of strings.
    if "permissions" not in cols(conn, "users"):
        conn.execute("ALTER TABLE users ADD COLUMN permissions TEXT;")
        # Backfill existing "user" accounts with the permissions that were
        # OPEN to everyone before this update, so nobody's access silently
        # shrinks. Import/Export/Purge were already Super-Admin-only before,
        # so they stay off by default — Super Admin can opt individual users
        # in from Manage Users.
        import json as _json
        default_perms = _json.dumps(["can_delete", "can_batch_renew", "can_location_check"])
        conn.execute("UPDATE users SET permissions = ? WHERE role != 'super_admin' AND permissions IS NULL", (default_perms,))
    # activity log — records who did what to which license, for history tracking
    conn.execute("""CREATE TABLE IF NOT EXISTS activity_log (
        id         INTEGER PRIMARY KEY AUTOINCREMENT,
        ts         TEXT NOT NULL,
        username   TEXT,
        action     TEXT NOT NULL,
        license_id INTEGER,
        license_no TEXT,
        detail     TEXT
    );""")
    conn.execute("""CREATE TABLE IF NOT EXISTS schema_meta (
        key TEXT PRIMARY KEY, value TEXT
    );""")
    # v2: New/Save/Save-as-New and Print/Preview/Open File become individually
    # controllable too. They were open to everyone before — grandfather every
    # existing "user" account in so nobody's access silently disappears; the
    # Super Admin can then uncheck them per-person from Manage Users if wanted.
    migrated_v2 = conn.execute("SELECT value FROM schema_meta WHERE key='perms_v2'").fetchone()
    if not migrated_v2:
        import json as _json
        rows = conn.execute("SELECT username, permissions FROM users WHERE role != 'super_admin'").fetchall()
        for r in rows:
            try:
                existing = _json.loads(r["permissions"]) if r["permissions"] else []
            except Exception:
                existing = []
            changed = False
            for p in ("can_create", "can_edit", "can_print"):
                if p not in existing:
                    existing.append(p); changed = True
            if changed:
                conn.execute("UPDATE users SET permissions = ? WHERE username = ?", (_json.dumps(existing), r["username"]))
        conn.execute("INSERT OR REPLACE INTO schema_meta (key, value) VALUES ('perms_v2', '1')")
    conn.commit()
    conn.close()

ensure_schema()

ALL_PERMISSIONS = ["can_create", "can_edit", "can_delete", "can_purge", "can_print",
                    "can_import", "can_export", "can_batch_renew", "can_location_check"]
PERMISSION_LABELS = {
    "can_create":         "Add new records (New / Save as New)",
    "can_edit":           "Save changes (overwrite existing record)",
    "can_delete":         "Delete records (move to Trash)",
    "can_purge":          "Permanently delete (empty Trash / purge)",
    "can_print":          "Print, Print Preview, and Open File",
    "can_import":         "Import records from Excel",
    "can_export":         "Export records to Excel",
    "can_batch_renew":    "Batch Renew",
    "can_location_check": "Location Check tool",
}

def _current_user_info(request):
    token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
    return _get_token_info(token)

def _user_permissions(username):
    import json as _json
    conn = get_conn()
    row = conn.execute("SELECT permissions FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()
    if not row or not row["permissions"]:
        return []
    try:
        return _json.loads(row["permissions"])
    except Exception:
        return []

def require_permission(request, perm):
    """Super Admin always passes. A 'user' account needs this specific
    permission checked in Manage Users, or the action is refused."""
    role = role_for(request)
    if role == "super_admin":
        return
    info = _current_user_info(request)
    if not info:
        raise HTTPException(401, "Not authenticated")
    if perm not in _user_permissions(info["username"]):
        label = PERMISSION_LABELS.get(perm, perm)
        raise HTTPException(403, f"You don't have permission for this ({label}). Ask your Super Admin to enable it for your account.")

def log_activity(request, action, license_id=None, license_no=None, detail=None):
    """Record an entry in the activity log. Never lets a logging problem
    break the actual request — failures here are swallowed silently."""
    import datetime
    try:
        info = _current_user_info(request) if request is not None else None
        username = info["username"] if info else None
        conn = get_conn()
        conn.execute(
            "INSERT INTO activity_log (ts, username, action, license_id, license_no, detail) VALUES (?,?,?,?,?,?)",
            (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), username, action, license_id, license_no, detail),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

def _diff_summary(old_row, new_payload, max_fields=8):
    """Build a short human-readable list of which fields changed and how,
    for the activity log entry on an edit."""
    changes = []
    for k, new_v in new_payload.items():
        if k in PROTECTED:
            continue
        old_v = old_row.get(k) if old_row else None
        old_norm = "" if old_v is None else str(old_v)
        new_norm = "" if new_v is None else str(new_v)
        if old_norm != new_norm:
            old_disp = old_norm if old_norm else "(blank)"
            new_disp = new_norm if new_norm else "(blank)"
            changes.append(f"{k}: '{old_disp}' → '{new_disp}'")
    if not changes:
        return "No field changes"
    shown = changes[:max_fields]
    more = len(changes) - len(shown)
    text = "; ".join(shown)
    if more > 0:
        text += f"; and {more} more field(s)"
    return text


def _effective_permissions(username, role):
    """Super Admin implicitly has every permission; a 'user' account has
    exactly what's been checked for them in Manage Users."""
    if role == "super_admin":
        return list(ALL_PERMISSIONS)
    return _user_permissions(username)


# Password login lockout — same shape as the PIN lockout below, so a
# brute-force attempt against the full password is throttled exactly like
# repeated wrong PINs already are, instead of being allowed unlimited tries.
LOGIN_ATTEMPTS = {}   # username -> {"count": int, "locked_until": epoch_seconds or None}
LOGIN_LOCKOUT_SECONDS = 5 * 60
LOGIN_MAX_ATTEMPTS = 5

@app.post("/api/login")
def login(data: dict = Body(...)):
    import time
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""

    state = LOGIN_ATTEMPTS.get(username, {"count": 0, "locked_until": None})
    now = time.time()
    if state["locked_until"] and now < state["locked_until"]:
        wait_min = max(1, int((state["locked_until"] - now) / 60) + 1)
        raise HTTPException(429, f"Too many failed sign-in attempts. Try again in about {wait_min} minute(s).")

    conn = get_conn()
    row = conn.execute("SELECT salt, pwhash, role FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    ok = bool(row and hash_pw(password, row["salt"]) == row["pwhash"])
    if not ok:
        state["count"] += 1
        if state["count"] >= LOGIN_MAX_ATTEMPTS:
            state["locked_until"] = now + LOGIN_LOCKOUT_SECONDS
            state["count"] = 0
        LOGIN_ATTEMPTS[username] = state
        raise HTTPException(401, "Invalid username or password")

    LOGIN_ATTEMPTS.pop(username, None)
    role = row["role"] or "admin"
    token = secrets.token_urlsafe(32)
    TOKENS[token] = _new_token_entry(username, role)
    return {"ok": True, "token": token, "username": username, "role": role,
            "permissions": _effective_permissions(username, role)}


# ---- PIN sign-in: a quick 4-digit alternative to the full password ----
PIN_ATTEMPTS = {}   # username -> {"count": int, "locked_until": epoch_seconds or None}
PIN_LOCKOUT_SECONDS = 5 * 60
PIN_MAX_ATTEMPTS = 5

@app.post("/api/pin-login")
def pin_login(data: dict = Body(...)):
    import time
    username = (data.get("username") or "").strip()
    pin = (data.get("pin") or "").strip()

    state = PIN_ATTEMPTS.get(username, {"count": 0, "locked_until": None})
    now = time.time()
    if state["locked_until"] and now < state["locked_until"]:
        wait_min = max(1, int((state["locked_until"] - now) / 60) + 1)
        raise HTTPException(429, f"Too many wrong PIN attempts. Try again in about {wait_min} minute(s), or sign in with the password instead.")

    conn = get_conn()
    row = conn.execute("SELECT pin_salt, pin_hash, role FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    ok = bool(row and row["pin_hash"] and hash_pw(pin, row["pin_salt"]) == row["pin_hash"])
    if not ok:
        state["count"] += 1
        if state["count"] >= PIN_MAX_ATTEMPTS:
            state["locked_until"] = now + PIN_LOCKOUT_SECONDS
            state["count"] = 0
        PIN_ATTEMPTS[username] = state
        if not row or not row["pin_hash"]:
            raise HTTPException(401, "No PIN is set for this account. Sign in with the password instead.")
        raise HTTPException(401, "Incorrect PIN")

    PIN_ATTEMPTS.pop(username, None)
    role = row["role"] or "admin"
    token = secrets.token_urlsafe(32)
    TOKENS[token] = _new_token_entry(username, role)
    return {"ok": True, "token": token, "username": username, "role": role,
            "permissions": _effective_permissions(username, role)}


@app.post("/api/set-my-pin")
def set_my_pin(data: dict = Body(...), request: Request = None):
    token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
    info = TOKENS.get(token)
    if not info:
        raise HTTPException(401, "Not authenticated")
    pin = (data.get("pin") or "").strip()
    if not (pin.isdigit() and len(pin) == 4):
        raise HTTPException(400, "PIN must be exactly 4 digits")
    salt = secrets.token_bytes(16)
    pin_hash = hash_pw(pin, salt)
    conn = get_conn()
    conn.execute("UPDATE users SET pin_salt = ?, pin_hash = ? WHERE username = ?",
                 (salt, pin_hash, info["username"]))
    conn.commit(); conn.close()
    return {"ok": True}


@app.post("/api/clear-my-pin")
def clear_my_pin(request: Request):
    token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
    info = TOKENS.get(token)
    if not info:
        raise HTTPException(401, "Not authenticated")
    conn = get_conn()
    conn.execute("UPDATE users SET pin_salt = NULL, pin_hash = NULL WHERE username = ?", (info["username"],))
    conn.commit(); conn.close()
    return {"ok": True}


@app.post("/api/logout")
def logout(request: Request):
    token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
    TOKENS.pop(token, None)
    return {"ok": True}


@app.get("/api/auth-status")
def auth_status():
    return {"has_account": any_user_exists()}


@app.get("/api/my-permissions")
def my_permissions(request: Request):
    """
    What the currently signed-in user is allowed to do — used by the
    frontend to hide/disable buttons for features they don't have, not just
    block the request after the fact.
    """
    info = _current_user_info(request)
    if not info:
        raise HTTPException(401, "Not authenticated")
    return {"role": info["role"], "permissions": _effective_permissions(info["username"], info["role"])}


@app.post("/api/setup-admin")
def setup_admin(data: dict = Body(...)):
    """
    Create the FIRST account — the Super Admin — from the app itself.
    Only works once: if any account already exists, this is refused (use the
    Manage Users panel, signed in as Super Admin, to add more accounts).
    """
    if any_user_exists():
        raise HTTPException(400, "An account already exists. Sign in, or ask the Super Admin to add you as a user.")
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username:
        raise HTTPException(400, "Username is required")
    if len(password) < 4:
        raise HTTPException(400, "Password must be at least 4 characters")

    salt = secrets.token_bytes(16)
    pwhash = hash_pw(password, salt)
    conn = get_conn()
    conn.execute(
        "INSERT INTO users (username, salt, pwhash, role) VALUES (?, ?, ?, 'super_admin')",
        (username, salt, pwhash))
    conn.commit(); conn.close()

    token = secrets.token_urlsafe(32)
    TOKENS[token] = _new_token_entry(username, "super_admin")
    return {"ok": True, "token": token, "username": username, "role": "super_admin",
            "permissions": list(ALL_PERMISSIONS)}


def require_super_admin(request: Request):
    if role_for(request) != "super_admin":
        raise HTTPException(403, "Only the Super Admin can manage users")


# ---------------------------------------------------------------- SETTINGS
# A simple shared file both the app and desktop.py's backup monitor read —
# so changing a setting here takes effect immediately, no code editing needed.
SETTINGS_FILE = "settings.json"
DEFAULT_SETTINGS = {"gdrive_backup_folder": r"G:\My Drive\NTC-Backups"}


def _load_settings():
    import json
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, SETTINGS_FILE)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            merged = dict(DEFAULT_SETTINGS)
            merged.update(data)
            return merged
        except Exception:
            pass
    return dict(DEFAULT_SETTINGS)


def _save_settings(data):
    import json
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, SETTINGS_FILE)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


@app.get("/api/settings")
def get_settings(request: Request):
    require_super_admin(request)
    return _load_settings()


@app.post("/api/settings")
def update_settings(data: dict = Body(...), request: Request = None):
    require_super_admin(request)
    current = _load_settings()
    if "gdrive_backup_folder" in data:
        current["gdrive_backup_folder"] = str(data["gdrive_backup_folder"]).strip()
    _save_settings(current)
    return {"ok": True, "settings": current}


@app.get("/api/backup-status")
def get_backup_status():
    """
    What the background backup monitor (in desktop.py) is doing right now —
    read by the small status badge in the app header. Available to anyone
    signed in, not just Super Admin, since it's just informational.
    """
    import json
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, "backup_status.json")
    if not os.path.exists(path):
        return {"state": "unknown"}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"state": "unknown"}


def _write_backup_status(**fields):
    import json, datetime
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, "backup_status.json")
    data = {}
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = {}
    data.update(fields)
    data["updated_at"] = datetime.datetime.now().isoformat(timespec="seconds")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


@app.post("/api/backup/run-now")
def run_backup_now(request: Request):
    """
    Manual 'push backup now' — copies the app folder (including telco.db) to
    the same single backup folder the automatic daily backup uses, right
    away, on demand. Doesn't change the automatic schedule at all: today's
    date still gets recorded, so the background monitor simply sees today's
    backup as already done and won't repeat it — nothing about the auto
    backup's timing or behavior is altered by using this button.
    telco.db is snapshotted via SQLite's own backup API (see
    _snapshot_sqlite_db) for a guaranteed-consistent copy, rather than a
    plain file copy that could in theory land mid-write.
    """
    require_super_admin(request)
    import shutil, datetime, sqlite3

    here = os.path.dirname(os.path.abspath(__file__))
    today = datetime.date.today().strftime("%Y-%m-%d")
    settings = _load_settings()
    gdrive_folder = settings.get("gdrive_backup_folder") or DEFAULT_SETTINGS["gdrive_backup_folder"]
    local_fallback = r"C:\NTC-App-Backups"
    dest_root = gdrive_folder if os.path.isdir(os.path.dirname(gdrive_folder) or "C:\\") else local_fallback
    dest = os.path.join(dest_root, "NTC-App-Backup")   # same single fixed folder as the auto backup

    try:
        _write_backup_status(state="backing_up")
        os.makedirs(dest_root, exist_ok=True)
        shutil.copytree(
            here, dest, dirs_exist_ok=True,
            ignore=shutil.ignore_patterns("venv", "__pycache__", "*.pyc", "telco.db"),
        )
        src_db = os.path.join(here, "telco.db")
        if os.path.exists(src_db):
            src_conn = sqlite3.connect(src_db)
            dest_conn = sqlite3.connect(os.path.join(dest, "telco.db"))
            try:
                with dest_conn:
                    src_conn.backup(dest_conn)
            finally:
                dest_conn.close(); src_conn.close()
        _write_backup_status(state="done", last_backup=today, last_backup_path=dest)
        return {"ok": True, "path": dest}
    except Exception as e:
        _write_backup_status(state="error", last_error=str(e))
        raise HTTPException(500, f"Backup failed: {e}")


@app.get("/api/users")
def list_users(request: Request):
    require_super_admin(request)
    import json as _json
    conn = get_conn()
    rows = conn.execute("SELECT username, role, pin_hash, permissions FROM users ORDER BY "
                         "CASE role WHEN 'super_admin' THEN 0 ELSE 1 END, username").fetchall()
    conn.close()
    out = []
    for r in rows:
        try:
            perms = _json.loads(r["permissions"]) if r["permissions"] else []
        except Exception:
            perms = []
        out.append({"username": r["username"], "role": r["role"],
                     "has_pin": r["pin_hash"] is not None, "permissions": perms})
    return {"users": out}


@app.get("/api/permissions-catalog")
def permissions_catalog(request: Request):
    """The full list of toggleable permissions, with labels — used to build
    the checklist in Manage Users."""
    require_super_admin(request)
    return {"permissions": [{"key": k, "label": PERMISSION_LABELS[k]} for k in ALL_PERMISSIONS]}


@app.get("/api/activity-log")
def get_activity_log(license_no: str = "", username: str = "", action: str = "", limit: int = 300, request: Request = None):
    """
    History tracking: who did what, and to which record. Super Admin only.
    Optional filters: license_no (partial match), username (exact),
    action (exact: create/edit/delete/restore/purge/import/batch_renew).
    """
    require_super_admin(request)
    where, params = [], []
    if license_no:
        where.append("license_no LIKE ?"); params.append(f"%{license_no}%")
    if username:
        where.append("username = ?"); params.append(username)
    if action:
        where.append("action = ?"); params.append(action)
    clause = ("WHERE " + " AND ".join(where)) if where else ""
    conn = get_conn()
    rows = conn.execute(
        f"SELECT id, ts, username, action, license_id, license_no, detail "
        f"FROM activity_log {clause} ORDER BY id DESC LIMIT ?",
        params + [max(1, min(limit, 2000))],
    ).fetchall()
    total = conn.execute(f"SELECT COUNT(*) FROM activity_log {clause}", params).fetchone()[0]
    conn.close()
    return {"count": len(rows), "total": total, "results": [dict(r) for r in rows]}


@app.post("/api/users")
def add_user(data: dict = Body(...), request: Request = None):
    require_super_admin(request)
    import json as _json
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    role = data.get("role") or "user"
    pin = (data.get("pin") or "").strip()   # optional
    perms = data.get("permissions")
    if role not in ("super_admin", "user"):
        role = "user"
    if not username:
        raise HTTPException(400, "Username is required")
    if pin and not (pin.isdigit() and len(pin) == 4):
        raise HTTPException(400, "PIN must be exactly 4 digits (or left blank)")
    perms_json = None
    clean_perms = None
    if perms is not None:
        clean_perms = [p for p in perms if p in ALL_PERMISSIONS]
        perms_json = _json.dumps(clean_perms)

    conn = get_conn()
    existing_row = conn.execute("SELECT role, permissions FROM users WHERE username = ?", (username,)).fetchone()
    exists = existing_row is not None

    # Snapshot the before-state so we can write a clear audit trail entry —
    # who changed what, on an account that can print, delete, and export
    # real licensing records, is exactly the kind of thing that should be
    # traceable later.
    old_role = existing_row["role"] if existing_row else None
    try:
        old_perms = set(_json.loads(existing_row["permissions"])) if existing_row and existing_row["permissions"] else set()
    except Exception:
        old_perms = set()

    if exists:
        # Editing an existing account: password and PIN are each optional —
        # leave either blank to keep it unchanged. Role and permissions
        # always update to whatever was submitted.
        if password and len(password) < 4:
            raise HTTPException(400, "Password must be at least 4 characters (or leave blank to keep it unchanged)")
        sets, params = ["role=?"], [role]
        if perms_json is not None:
            sets += ["permissions=?"]; params += [perms_json]
        pw_changed = False
        pin_changed = False
        if password:
            salt = secrets.token_bytes(16)
            sets += ["salt=?", "pwhash=?"]
            params += [salt, hash_pw(password, salt)]
            pw_changed = True
        if pin:
            pin_salt = secrets.token_bytes(16)
            sets += ["pin_salt=?", "pin_hash=?"]
            params += [pin_salt, hash_pw(pin, pin_salt)]
            pin_changed = True
        params.append(username)
        conn.execute(f"UPDATE users SET {','.join(sets)} WHERE username=?", params)
        action = "updated"

        # A role change takes effect immediately for anyone already signed
        # in, instead of waiting for their next login — otherwise a
        # just-demoted account would keep acting as Super Admin (or a
        # just-promoted one would stay stuck as a limited user) for the
        # rest of that session.
        if old_role != role:
            for tok_info in TOKENS.values():
                if tok_info.get("username") == username:
                    tok_info["role"] = role

        change_bits = []
        if old_role != role:
            change_bits.append(f"role: '{old_role}' -> '{role}'")
        if clean_perms is not None:
            new_perms = set(clean_perms)
            added = sorted(new_perms - old_perms)
            removed = sorted(old_perms - new_perms)
            if added:
                change_bits.append(f"permissions granted: {', '.join(added)}")
            if removed:
                change_bits.append(f"permissions removed: {', '.join(removed)}")
        if pw_changed:
            change_bits.append("password changed")
        if pin_changed:
            change_bits.append("PIN changed")
        detail = f"Account '{username}' updated" + ((" — " + "; ".join(change_bits)) if change_bits else " (no changes)")
        log_activity(request, "user_management", detail=detail)
    else:
        if len(password) < 4:
            conn.close()
            raise HTTPException(400, "Password must be at least 4 characters")
        salt = secrets.token_bytes(16)
        pwhash = hash_pw(password, salt)
        pin_salt = secrets.token_bytes(16) if pin else None
        pin_hash = hash_pw(pin, pin_salt) if pin else None
        conn.execute("INSERT INTO users (username, salt, pwhash, role, pin_salt, pin_hash, permissions) VALUES (?,?,?,?,?,?,?)",
                     (username, salt, pwhash, role, pin_salt, pin_hash, perms_json if perms_json is not None else _json.dumps([])))
        action = "added"
        perm_note = f" with permissions: {', '.join(clean_perms)}" if clean_perms else ""
        log_activity(request, "user_management", detail=f"Account '{username}' created as role '{role}'{perm_note}")
    conn.commit(); conn.close()
    return {"ok": True, "username": username, "role": role, "action": action}


@app.post("/api/users/{username}/clear-pin")
def clear_user_pin(username: str, request: Request):
    require_super_admin(request)
    conn = get_conn()
    cur = conn.execute("UPDATE users SET pin_salt = NULL, pin_hash = NULL WHERE username = ?", (username,))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "User not found")
    return {"ok": True}


@app.delete("/api/users/{username}")
def remove_user(username: str, request: Request):
    require_super_admin(request)
    me = None
    token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
    info = TOKENS.get(token)
    if info:
        me = info["username"]
    if username == me:
        raise HTTPException(400, "You can't remove your own account while signed in as it")
    conn = get_conn()
    cur = conn.execute("DELETE FROM users WHERE username = ?", (username,))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "User not found")
    # invalidate any active sessions for that user
    dead = [t for t, v in TOKENS.items() if v["username"] == username]
    for t in dead: TOKENS.pop(t, None)
    log_activity(request, "user_management", detail=f"Account '{username}' removed")
    return {"ok": True, "removed": username}


@app.middleware("http")
async def auth_guard(request: Request, call_next):
    path = request.url.path
    open_paths = {"/api/login", "/api/pin-login", "/api/logout", "/api/auth-status", "/api/setup-admin"}
    if path.startswith("/api") and path not in open_paths:
        token = request.headers.get("authorization", "").replace("Bearer ", "").strip()
        if token not in TOKENS:
            return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    return await call_next(request)


@app.on_event("shutdown")
def _close_printer():
    try:
        import print_engine
        print_engine.shutdown()
    except Exception:
        pass
    try:
        import print_engine_word
        print_engine_word.shutdown()
    except Exception:
        pass

# ---------------------------------------------------------------- root (GUI)
@app.get("/", response_class=HTMLResponse)
def root():
    if os.path.exists("index.html"):
        return FileResponse("index.html")
    return HTMLResponse(
        "<h2>Backend running.</h2><p>Put index.html in this folder, "
        "or visit <a href='/docs'>/docs</a>.</p>")

# ---------------------------------------------------------------- meta (for dropdowns)
@app.get("/api/meta")
def meta():
    conn = get_conn()
    def distinct(col):
        return [r[0] for r in conn.execute(
            f"SELECT DISTINCT {col} FROM licenses WHERE {col} IS NOT NULL "
            f"AND {col} <> '' ORDER BY {col}")]
    out = {
        "provinces": distinct("province"),
        "licensees": distinct("licensee"),
        "statuses":  distinct("license_status"),
        "classes":   distinct("class_of_station"),
        "techs":     distinct("tech"),
        "columns":   cols(conn, "licenses"),
    }
    conn.close()
    return out

# ---------------------------------------------------------------- history chain
@app.get("/api/licenses/{lic_id}/history")
def history(lic_id: int):
    conn = get_conn()
    chain = []
    # walk up the renewed_from links to the earliest cycle
    seen = set()
    cur_id = lic_id
    while cur_id and cur_id not in seen:
        seen.add(cur_id)
        row = conn.execute(
            "SELECT id, license_no, rsl_date, validity_from, validity_to, "
            "new_form_no, old_form_no, renewed_from FROM licenses WHERE id = ?",
            (cur_id,)).fetchone()
        if not row:
            break
        chain.append(dict(row))
        cur_id = row["renewed_from"]
    # also pull any newer cycles that point back to lic_id
    kids = conn.execute(
        "SELECT id, license_no, rsl_date, validity_from, validity_to, "
        "new_form_no, old_form_no, renewed_from FROM licenses "
        "WHERE renewed_from = ?", (lic_id,)).fetchall()
    conn.close()
    chain = list(reversed(chain))  # oldest first
    return {"chain": chain, "newer": [dict(k) for k in kids]}

# ---------------------------------------------------------------- stats
@app.get("/api/stats")
def stats():
    conn = get_conn()
    def grp(col):
        return {r[0]: r[1] for r in conn.execute(
            f"SELECT {col}, COUNT(*) FROM licenses WHERE deleted_at IS NULL "
            f"GROUP BY {col} ORDER BY 2 DESC")}
    out = {
        "total_licenses": conn.execute("SELECT COUNT(*) FROM licenses WHERE deleted_at IS NULL").fetchone()[0],
        "total_payments": conn.execute("SELECT COUNT(*) FROM payments").fetchone()[0],
        "trash_count":  conn.execute("SELECT COUNT(*) FROM licenses WHERE deleted_at IS NOT NULL").fetchone()[0],
        "by_province":  grp("province"),
        "by_licensee":  grp("licensee"),
        "by_status":    grp("license_status"),
    }
    conn.close()
    return out

# ---------------------------------------------------------------- list / search
@app.get("/api/licenses")
def list_licenses(
    q: str = Query("", description="search text"),
    province: str = "",
    licensee: str = "",
    status: str = "",
    limit: int = 50,
    offset: int = 0,
):
    conn = get_conn()
    where, params = ["deleted_at IS NULL"], []
    if q:
        like = f"%{q}%"
        where.append("(license_no LIKE ? OR site_no LIKE ? OR site_name LIKE ? OR brgy LIKE ? "
                     "OR town LIKE ? OR province LIKE ? OR licensee LIKE ? OR tech LIKE ? "
                     "OR or_no LIKE ? OR or_date LIKE ?)")
        params += [like] * 10
    if province:
        where.append("province = ?"); params.append(province)
    if licensee:
        where.append("licensee = ?"); params.append(licensee)
    if status:
        where.append("license_status = ?"); params.append(status)

    clause = ("WHERE " + " AND ".join(where)) if where else ""
    total = conn.execute(f"SELECT COUNT(*) FROM licenses {clause}", params).fetchone()[0]
    rows = conn.execute(
        f"SELECT id, license_no, licensee, site_name, brgy, town, province, "
        f"class_of_station, tech, validity_from, validity_to, license_status, status, "
        f"or_no, or_date, or_amount, created_at, updated_at "
        f"FROM licenses {clause} ORDER BY id LIMIT ? OFFSET ?",
        params + [limit, offset],
    ).fetchall()
    conn.close()
    return {"total": total, "count": len(rows), "limit": limit, "offset": offset,
            "results": [dict(r) for r in rows]}

# ---------------------------------------------------------------- get one (+payments)
@app.get("/api/licenses/{lic_id}")
def get_license(lic_id: int):
    conn = get_conn()
    row = conn.execute("SELECT * FROM licenses WHERE id = ?", (lic_id,)).fetchone()
    if not row:
        conn.close(); raise HTTPException(404, "License not found")
    pays = conn.execute(
        "SELECT * FROM payments WHERE license_id = ? ORDER BY year", (lic_id,)).fetchall()
    conn.close()
    return {"license": dict(row), "payments": [dict(p) for p in pays]}

# ---------------------------------------------------------------- create
@app.post("/api/licenses")
def create_license(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_create")
    conn = get_conn()
    payload = clean_payload(conn, "licenses", data)
    if not payload:
        conn.close(); raise HTTPException(400, "No valid fields supplied")
    fields = list(payload.keys())
    sql = (f"INSERT INTO licenses ({','.join(fields)}) "
           f"VALUES ({','.join('?' * len(fields))})")
    cur = conn.execute(sql, [payload[f] for f in fields])
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    log_activity(request, "create", license_id=new_id, license_no=payload.get("license_no"),
                 detail="New record created" + (f" (renewed/modified from #{payload['renewed_from']})" if payload.get("renewed_from") else ""))
    return {"ok": True, "id": new_id}

# ---------------------------------------------------------------- update
@app.put("/api/licenses/{lic_id}")
def update_license(lic_id: int, data: dict = Body(...), request: Request = None):
    require_permission(request, "can_edit")
    conn = get_conn()
    old_row = conn.execute("SELECT * FROM licenses WHERE id = ?", (lic_id,)).fetchone()
    if not old_row:
        conn.close(); raise HTTPException(404, "License not found")
    old_dict = dict(old_row)
    payload = clean_payload(conn, "licenses", data)
    if not payload:
        conn.close(); raise HTTPException(400, "No valid fields supplied")
    payload["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sets = ", ".join(f"{f} = ?" for f in payload)
    conn.execute(f"UPDATE licenses SET {sets} WHERE id = ?",
                 list(payload.values()) + [lic_id])
    conn.commit(); conn.close()
    log_activity(request, "edit", license_id=lic_id, license_no=old_dict.get("license_no"),
                 detail=_diff_summary(old_dict, payload))
    return {"ok": True, "id": lic_id}

# ---------------------------------------------------------------- delete (to Trash) + Trash bin
@app.delete("/api/licenses/{lic_id}")
def delete_license(lic_id: int, request: Request):
    """Soft delete: move the record to the Trash (recoverable)."""
    require_permission(request, "can_delete")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    row = conn.execute("SELECT license_no FROM licenses WHERE id = ?", (lic_id,)).fetchone()
    cur = conn.execute(
        "UPDATE licenses SET deleted_at = ? WHERE id = ? AND deleted_at IS NULL",
        (now, lic_id))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "License not found")
    log_activity(request, "delete", license_id=lic_id, license_no=row["license_no"] if row else None,
                 detail="Moved to Trash")
    return {"ok": True, "trashed": lic_id}


# ---------------------------------------------------------------- LOCATION CHECK (find likely-wrong coordinates)
import math as _math

# Loose sanity box for Region II, Philippines — anything outside this is
# almost certainly a data-entry error (wrong field, swapped digits, etc.)
REGION2_LAT_MIN, REGION2_LAT_MAX = 15.5, 21.5
REGION2_LNG_MIN, REGION2_LNG_MAX = 120.5, 122.8

def _haversine_km(lat1, lng1, lat2, lng2):
    R = 6371.0
    p1, p2 = _math.radians(lat1), _math.radians(lat2)
    dp = _math.radians(lat2 - lat1)
    dl = _math.radians(lng2 - lng1)
    a = _math.sin(dp/2)**2 + _math.cos(p1)*_math.cos(p2)*_math.sin(dl/2)**2
    return 2 * R * _math.asin(min(1, _math.sqrt(a)))

def _median(vals):
    s = sorted(vals)
    n = len(s)
    if n == 0: return None
    mid = n // 2
    return s[mid] if n % 2 else (s[mid-1] + s[mid]) / 2

@app.get("/api/location-check")
def location_check(radius_km: float = 5.0, request: Request = None):
    """
    Flags records whose coordinates look wrong:
      - missing: no usable coordinates at all
      - impossible: outside the loose Region II sanity box (swapped fields,
        stray digits, etc.)
      - outlier: further than `radius_km` from the median position of other
        stations in the same barangay (falls back to town, then province, if
        the barangay doesn't have enough neighbors to compare against)
    Nothing is changed — this only reports what to review.
    """
    require_permission(request, "can_location_check")
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, license_no, licensee, site_name, brgy, town, province, "
        "nlat_deg, nlat_min, nlat_sec, elong_deg, elong_min, elong_sec "
        "FROM licenses WHERE deleted_at IS NULL"
    ).fetchall()
    conn.close()

    recs = []
    for r in rows:
        lat = _dms_to_dec(r["nlat_deg"], r["nlat_min"], r["nlat_sec"]) or 0
        lng = _dms_to_dec(r["elong_deg"], r["elong_min"], r["elong_sec"]) or 0
        recs.append({
            "id": r["id"], "license_no": r["license_no"], "licensee": r["licensee"],
            "site_name": r["site_name"], "brgy": r["brgy"], "town": r["town"],
            "province": r["province"], "lat": lat, "lng": lng,
        })

    flagged = []

    # group keys, from most to least specific, for the "compare to neighbors" check
    def key_brgy(rc):    return (rc["province"] or "", rc["town"] or "", rc["brgy"] or "")
    def key_town(rc):    return (rc["province"] or "", rc["town"] or "")
    def key_province(rc):return (rc["province"] or "",)

    valid = [rc for rc in recs if rc["lat"] > 0 and rc["lng"] > 0
             and REGION2_LAT_MIN <= rc["lat"] <= REGION2_LAT_MAX
             and REGION2_LNG_MIN <= rc["lng"] <= REGION2_LNG_MAX]

    def group_medians(keyfn):
        groups = {}
        for rc in valid:
            groups.setdefault(keyfn(rc), []).append(rc)
        meds = {}
        for k, members in groups.items():
            if len(members) >= 3:   # need at least a few neighbors to trust a median
                meds[k] = (_median([m["lat"] for m in members]), _median([m["lng"] for m in members]), members)
        return meds

    brgy_meds    = group_medians(key_brgy)
    town_meds    = group_medians(key_town)
    province_meds= group_medians(key_province)

    MAX_NEIGHBORS = 40   # cap the neighbor list so provincial fallback groups stay light

    def find_group(rc):
        """Return group info + the actual neighbor stations for the tightest
        group with enough members to trust, or {} if nowhere has enough."""
        for keyfn, meds, label in ((key_brgy, brgy_meds, "barangay"), (key_town, town_meds, "town"), (key_province, province_meds, "province")):
            k = keyfn(rc)
            if k in meds:
                mlat, mlng, members = meds[k]
                others = [m for m in members if m["id"] != rc["id"]][:MAX_NEIGHBORS]
                return {
                    "group_lat": mlat, "group_lng": mlng, "group_basis": label, "group_n": len(members),
                    "group_points": [{"id": m["id"], "license_no": m["license_no"], "lat": m["lat"], "lng": m["lng"]} for m in others],
                }
        return {}

    for rc in recs:
        lat, lng = rc["lat"], rc["lng"]
        if lat <= 0 or lng <= 0:
            flagged.append({**rc, "reason": "missing", "detail": "No coordinates entered",
                             "distance_km": None, **find_group(rc)})
            continue
        if not (REGION2_LAT_MIN <= lat <= REGION2_LAT_MAX and REGION2_LNG_MIN <= lng <= REGION2_LNG_MAX):
            flagged.append({**rc, "reason": "impossible",
                             "detail": f"Coordinates fall outside Region II (lat {lat:.4f}, lng {lng:.4f}) — check for swapped or mistyped values",
                             "distance_km": None, **find_group(rc)})
            continue
        # compare to the tightest group that has enough neighbors
        group = find_group(rc)
        if not group:
            continue   # not enough neighbors anywhere to judge this one
        dist = _haversine_km(lat, lng, group["group_lat"], group["group_lng"])
        if dist > radius_km:
            flagged.append({**rc, "reason": "outlier",
                             "detail": f"{dist:.1f} km from the other {group['group_n']} station(s) in the same {group['group_basis']}",
                             "distance_km": round(dist, 1), **group})

    # worst first: impossible/missing before distance outliers, then farthest first
    order = {"impossible": 0, "missing": 0, "outlier": 1}
    flagged.sort(key=lambda f: (order.get(f["reason"], 2), -(f["distance_km"] or 0)))
    return {"count": len(flagged), "checked": len(recs), "results": flagged}


@app.get("/api/trash")
def list_trash():
    conn = get_conn()
    rows = conn.execute(
        "SELECT id, license_no, licensee, site_name, brgy, town, province, "
        "class_of_station, tech, license_status, or_no, deleted_at "
        "FROM licenses WHERE deleted_at IS NOT NULL ORDER BY deleted_at DESC").fetchall()
    conn.close()
    return {"count": len(rows), "results": [dict(r) for r in rows]}


@app.post("/api/licenses/{lic_id}/restore")
def restore_license(lic_id: int, request: Request):
    require_permission(request, "can_delete")
    conn = get_conn()
    row = conn.execute("SELECT license_no FROM licenses WHERE id = ?", (lic_id,)).fetchone()
    cur = conn.execute(
        "UPDATE licenses SET deleted_at = NULL WHERE id = ? AND deleted_at IS NOT NULL",
        (lic_id,))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "Record not in trash")
    log_activity(request, "restore", license_id=lic_id, license_no=row["license_no"] if row else None,
                 detail="Restored from Trash")
    return {"ok": True, "restored": lic_id}


@app.delete("/api/licenses/{lic_id}/purge")
def purge_license(lic_id: int, request: Request):
    """Permanently delete one trashed record (cannot be undone)."""
    require_permission(request, "can_purge")
    conn = get_conn()
    row = conn.execute("SELECT license_no FROM licenses WHERE id = ?", (lic_id,)).fetchone()
    cur = conn.execute(
        "DELETE FROM licenses WHERE id = ? AND deleted_at IS NOT NULL", (lic_id,))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "Record not in trash")
    log_activity(request, "purge", license_id=lic_id, license_no=row["license_no"] if row else None,
                 detail="Permanently deleted")
    return {"ok": True, "purged": lic_id}


@app.post("/api/trash/empty")
def empty_trash(request: Request):
    """Permanently delete everything in the Trash (cannot be undone)."""
    require_permission(request, "can_purge")
    conn = get_conn()
    cur = conn.execute("DELETE FROM licenses WHERE deleted_at IS NOT NULL")
    conn.commit(); n = cur.rowcount; conn.close()
    log_activity(request, "purge", detail=f"Trash emptied — {n} record(s) permanently deleted")
    return {"ok": True, "purged": n}


@app.post("/api/wipe-all")
def wipe_all(data: dict = Body(...), request: Request = None):
    """
    Delete EVERY record (licenses + payments). Super Admin only.
    Requires the exact confirmation phrase "DELETE ALL" to proceed.
    Always backs up telco.db first — this cannot be undone otherwise.
    """
    require_super_admin(request)
    import shutil, os, datetime
    if (data.get("confirm") or "").strip() != "DELETE ALL":
        raise HTTPException(400, 'Type "DELETE ALL" exactly to confirm.')

    backup = None
    try:
        if os.path.exists(DB):
            backup = f"telco_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy(DB, backup)
    except Exception as e:
        raise HTTPException(500, f"Could not create backup, aborted for safety: {e}")

    conn = get_conn()
    n_lic = conn.execute("SELECT COUNT(*) FROM licenses").fetchone()[0]
    conn.execute("DELETE FROM payments")
    conn.execute("DELETE FROM licenses")
    try:
        conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('licenses','payments')")
    except Exception:
        pass
    conn.commit(); conn.close()
    return {"ok": True, "deleted": n_lic, "backup": backup}

# ---------------------------------------------------------------- payments
# ---------------------------------------------------------------- BATCH RENEW
@app.post("/api/batch/renew")
def batch_renew(data: dict = Body(...), request: Request = None):
    """
    Renew many RSLs at once with one shared Official Receipt.
      data = {
        "source_ids": [12, 15, 33, ...],
        "shared":     {"or_no":..., "or_date":..., "or_amount":..., "rsl_date":...}
      }
    For each source it creates a new RENEWAL record (original kept as history):
      - carries the current New Form No. -> Old Form No., and RSL Date -> Old Date
      - applies the shared OR number / date / amount and the RSL date
      - clears the New Form No. and the validity dates (set per record before printing)
    Returns the new record ids in the same order.
    """
    require_permission(request, "can_batch_renew")
    source_ids = data.get("source_ids", [])
    shared     = data.get("shared", {}) or {}
    if not source_ids:
        raise HTTPException(400, "No stations selected")

    SHARED_KEYS = {"rsl_date", "or_no", "or_date", "or_amount"}
    shared = {k: v for k, v in shared.items() if k in SHARED_KEYS}

    conn = get_conn()
    writable = set(cols(conn, "licenses")) - PROTECTED
    new_ids = []
    try:
        for sid in source_ids:
            row = conn.execute("SELECT * FROM licenses WHERE id = ?", (sid,)).fetchone()
            if not row:
                continue
            src = dict(row)
            rec = {k: v for k, v in src.items() if k in writable}

            # history link + auto-carry old form info
            rec["renewed_from"] = sid
            rec["old_form_no"]  = src.get("new_form_no")
            rec["old_date"]     = src.get("rsl_date")
            rec["new_form_no"]  = None          # blank until the new form is issued

            # clear the validity dates (new term set per record before printing)
            rec["validity_from"] = None
            rec["validity_to"]   = None

            # shared batch data (one OR covers the batch)
            for k, v in shared.items():
                rec[k] = v

            rec["status"] = "RENEWAL"
            rec["license_status"] = "RENEWAL"

            fields = list(rec.keys())
            sql = (f"INSERT INTO licenses ({','.join(fields)}) "
                   f"VALUES ({','.join('?'*len(fields))})")
            cur = conn.execute(sql, [rec[f] for f in fields])
            new_ids.append(cur.lastrowid)
        conn.commit()
    finally:
        conn.close()
    log_activity(request, "batch_renew", detail=f"Batch renewed {len(new_ids)} record(s) from {len(source_ids)} selected")
    return {"ok": True, "renewed": len(new_ids), "new_ids": new_ids}


# ---------------------------------------------------------------- ANALYTICS
def _build_analytics_rows(conn):
    """
    One row per unique License No — duplicates/renewal history collapsed to
    the version with the latest Validity To date (that one represents the
    license's current state: its Tech, Channel, Location, and expiry).
    Status is computed by comparing that date to today.
    """
    import datetime
    today = datetime.date.today()

    rows = conn.execute("""
        SELECT id, license_no, tech, config, brgy, town, province,
               validity_to, license_status, status, licensee, class_of_station,
               site_no, site_name, elong_deg, elong_min, elong_sec, nlat_deg, nlat_min, nlat_sec
        FROM licenses
        WHERE deleted_at IS NULL AND license_no IS NOT NULL AND license_no <> ''
    """).fetchall()

    latest = {}   # license_no -> best row so far
    for r in rows:
        lic = r["license_no"]
        cur = latest.get(lic)
        this_date = r["validity_to"] or ""
        if cur is None or this_date > (cur["validity_to"] or ""):
            latest[lic] = r

    out = []
    for lic, r in latest.items():
        vto = r["validity_to"]
        expired = None
        if vto:
            try:
                d = datetime.datetime.strptime(str(vto)[:10], "%Y-%m-%d").date()
                expired = d < today
            except ValueError:
                expired = None
        out.append({
            "id": r["id"],
            "license_no": lic,
            "licensee": r["licensee"],
            "tech": r["tech"],
            "config": r["config"],
            "class_of_station": r["class_of_station"],
            "site_no": r["site_no"],
            "site_name": r["site_name"],
            "brgy": r["brgy"],
            "town": r["town"],
            "province": r["province"],
            "validity_to": vto,
            "status": ("Expired" if expired else "Renewed") if expired is not None else "Unknown",
            "elong_deg": r["elong_deg"], "elong_min": r["elong_min"], "elong_sec": r["elong_sec"],
            "nlat_deg": r["nlat_deg"], "nlat_min": r["nlat_min"], "nlat_sec": r["nlat_sec"],
        })
    out.sort(key=lambda x: (x["province"] or "", x["town"] or "", x["license_no"] or ""))
    return out


PIVOT_DIMS = {
    "tech": "tech",
    "province": "province",
    "town": "town",
    "brgy": "brgy",
}
PIVOT_LABELS = {"tech": "Technology", "province": "Province", "town": "Municipality", "brgy": "Barangay"}

def _pivot_data(conn, x_dim, y_dim, licensee=None):
    if x_dim not in PIVOT_DIMS or y_dim not in PIVOT_DIMS:
        raise HTTPException(400, "Invalid axis — choose from tech, province, town, brgy")
    rows = _build_analytics_rows(conn)
    if licensee:
        rows = [r for r in rows if (r["licensee"] or "") == licensee]
    xk, yk = PIVOT_DIMS[x_dim], PIVOT_DIMS[y_dim]

    x_vals = sorted(set((r[xk] or f"(no {PIVOT_LABELS[x_dim].lower()})") for r in rows))
    grid = {}
    for r in rows:
        yv = r[yk] or f"(no {PIVOT_LABELS[y_dim].lower()})"
        xv = r[xk] or f"(no {PIVOT_LABELS[x_dim].lower()})"
        grid.setdefault(yv, {})
        grid[yv][xv] = grid[yv].get(xv, 0) + 1

    pivot_rows = []
    for yv, counts in sorted(grid.items()):
        pivot_rows.append({"y": yv, "counts": counts, "total": sum(counts.values())})
    col_totals = {xv: 0 for xv in x_vals}
    for r in pivot_rows:
        for xv, n in r["counts"].items():
            if xv in col_totals:
                col_totals[xv] += n
    grand_total = sum(col_totals.values())
    return {
        "x_dim": x_dim, "y_dim": y_dim,
        "x_label": PIVOT_LABELS[x_dim], "y_label": PIVOT_LABELS[y_dim],
        "x_vals": x_vals, "rows": pivot_rows,
        "col_totals": col_totals, "grand_total": grand_total,
    }


@app.get("/api/analytics/pivot")
def get_analytics_pivot(x: str = "tech", y: str = "province", licensee: str = ""):
    """
    Flexible cross-tab: pick any two dimensions (tech, province, town, brgy)
    for the X and Y axes. Cells = count of unique licenses (one-per-license
    rule, same as /api/analytics). Optionally restrict to one carrier.
    """
    conn = get_conn()
    try:
        return _pivot_data(conn, x, y, licensee or None)
    finally:
        conn.close()


@app.post("/api/analytics/pivot/export")
def export_analytics_pivot(x: str = "tech", y: str = "province", licensee: str = "", request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    conn = get_conn()
    try:
        p = _pivot_data(conn, x, y, licensee or None)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "pivot"
    if licensee:
        ws.append([f"Carrier: {licensee}"])
        ws.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, italic=True)
        ws.append([])
    headers = [p["y_label"]] + p["x_vals"] + ["Total"]
    ws.append(headers)
    header_row = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).font = openpyxl.styles.Font(bold=True)

    for r in p["rows"]:
        row = [r["y"]] + [r["counts"].get(xv, 0) for xv in p["x_vals"]] + [r["total"]]
        ws.append(row)

    total_row = ["TOTAL"] + [p["col_totals"].get(xv, 0) for xv in p["x_vals"]] + [p["grand_total"]]
    ws.append(total_row)
    for c in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=c).font = openpyxl.styles.Font(bold=True)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate
    ws.column_dimensions["A"].width = 22
    for i in range(len(p["x_vals"]) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(2 + i)].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_Pivot_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _dms_to_dec(deg, mins, secs):
    try:
        return (float(deg or 0) + float(mins or 0)/60 + float(secs or 0)/3600)
    except (TypeError, ValueError):
        return None


def _filter_by_y(rows, y_dim, values):
    yk = PIVOT_DIMS.get(y_dim)
    if not yk:
        raise HTTPException(400, "Invalid axis — choose from tech, province, town, brgy")
    wanted = set(values)
    out = []
    for r in rows:
        yv = r[yk] or f"(no {PIVOT_LABELS[y_dim].lower()})"
        if yv in wanted:
            out.append(r)
    return out


@app.get("/api/analytics/pivot/selected")
def get_pivot_selected(y: str, values: str, licensee: str = ""):
    """Records whose Y-axis value is in the given list (comma-separated) —
    used for 'export selected' and 'pin selected to map' from the pivot table."""
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()
    if licensee:
        rows = [r for r in rows if (r["licensee"] or "") == licensee]
    vals = [v for v in values.split(",") if v != ""]
    filtered = _filter_by_y(rows, y, vals)
    # attach decimal coordinates for easy map pinning
    for r in filtered:
        r["lat"] = _dms_to_dec(r["nlat_deg"], r["nlat_min"], r["nlat_sec"])
        r["lng"] = _dms_to_dec(r["elong_deg"], r["elong_min"], r["elong_sec"])
    return {"count": len(filtered), "rows": filtered}


@app.post("/api/analytics/pivot/selected/export")
def export_pivot_selected(y: str, values: str, licensee: str = "", request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()
    if licensee:
        rows = [r for r in rows if (r["licensee"] or "") == licensee]
    vals = [v for v in values.split(",") if v != ""]
    filtered = _filter_by_y(rows, y, vals)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "selected"
    headers = ["License No","Licensee","Tech","Channel/Config","Class of Station",
               "Site Name","Barangay","Municipality","Province","Validity To","Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for r in filtered:
        ws.append([r["license_no"], r["licensee"], r["tech"], r["config"], r["class_of_station"],
                   r["site_name"], r["brgy"], r["town"], r["province"], r["validity_to"], r["status"]])
    ws.freeze_panes = "A2"
    for i, w in enumerate([18,26,12,14,14,20,16,16,14,12,10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_Selected_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/recent")
def get_recent(limit: int = 50):
    """The most recently added/saved records, newest first — for the
    'Recently Added' screen, so nothing gets lost after a busy day of entry."""
    conn = get_conn()
    try:
        rows = conn.execute("""
            SELECT id, license_no, licensee, site_name, brgy, town, province,
                   status, tech, created_at, updated_at
            FROM licenses
            WHERE deleted_at IS NULL
            ORDER BY created_at DESC
            LIMIT ?
        """, (limit,)).fetchall()
    finally:
        conn.close()
    return {"records": [dict(r) for r in rows]}


@app.get("/api/or-batches")
def get_or_batches():
    """
    Groups every record by its (OR No., OR Date) pair — since Batch Renew
    writes the exact same OR info to every station in a batch, this pair
    naturally identifies "which stations were renewed together on this
    receipt." Always computed live from the database, so it reflects every
    batch renewal automatically — nothing extra to keep in sync.
    """
    conn = get_conn()
    try:
        rows = conn.execute("""
            SELECT id, license_no, licensee, site_name, brgy, town, province,
                   or_no, or_date, or_amount, rsl_date, status, tech,
                   elong_deg, elong_min, elong_sec, nlat_deg, nlat_min, nlat_sec
            FROM licenses
            WHERE deleted_at IS NULL AND or_no IS NOT NULL AND or_no <> ''
                  AND or_date IS NOT NULL AND or_date <> ''
            ORDER BY or_date DESC, or_no
        """).fetchall()
    finally:
        conn.close()

    groups = {}
    order = []
    for r in rows:
        key = (r["or_no"], r["or_date"])
        if key not in groups:
            groups[key] = {
                "or_no": r["or_no"], "or_date": r["or_date"],
                "or_amount": r["or_amount"], "records": [],
            }
            order.append(key)
        groups[key]["records"].append({
            "id": r["id"], "license_no": r["license_no"], "licensee": r["licensee"],
            "site_name": r["site_name"], "brgy": r["brgy"], "town": r["town"],
            "province": r["province"], "rsl_date": r["rsl_date"], "status": r["status"], "tech": r["tech"],
            "elong_deg": r["elong_deg"], "elong_min": r["elong_min"], "elong_sec": r["elong_sec"],
            "nlat_deg": r["nlat_deg"], "nlat_min": r["nlat_min"], "nlat_sec": r["nlat_sec"],
        })

    batches = [groups[k] for k in order]
    for b in batches:
        b["count"] = len(b["records"])
    return {"total_batches": len(batches), "batches": batches}


@app.post("/api/or-batches/export")
def export_or_batches(request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    data = get_or_batches()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "or_batches"
    headers = ["OR No.", "OR Date", "OR Amount", "# Records", "License No.", "Licensee",
               "Site Name", "Barangay", "Municipality", "Province", "RSL Date", "Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for b in data["batches"]:
        for r in b["records"]:
            ws.append([b["or_no"], b["or_date"], b["or_amount"], b["count"],
                       r["license_no"], r["licensee"], r["site_name"], r["brgy"],
                       r["town"], r["province"], r["rsl_date"], r["status"]])
    ws.freeze_panes = "A2"
    for i, w in enumerate([16,12,14,10,18,24,18,16,16,14,12,10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"OR_Batches_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/analytics")
def get_analytics():
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()
    expired_n = sum(1 for r in rows if r["status"] == "Expired")
    renewed_n = sum(1 for r in rows if r["status"] == "Renewed")
    return {"total_unique_licenses": len(rows), "expired": expired_n, "renewed": renewed_n, "rows": rows}


@app.post("/api/analytics/export")
def export_analytics(request: Request = None):
    require_permission(request, "can_export")
    import io, openpyxl, datetime
    conn = get_conn()
    try:
        rows = _build_analytics_rows(conn)
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "analytics"
    headers = ["License No","Licensee","Tech","Channel/Config","Class of Station",
               "Barangay","Municipality","Province","Validity To","Status"]
    ws.append(headers)
    for c in range(1, len(headers)+1):
        ws.cell(row=1, column=c).font = openpyxl.styles.Font(bold=True)
    for r in rows:
        ws.append([r["license_no"], r["licensee"], r["tech"], r["config"], r["class_of_station"],
                   r["brgy"], r["town"], r["province"], r["validity_to"], r["status"]])
    ws.freeze_panes = "A2"
    for i, w in enumerate([18,26,12,14,14,16,16,14,12,10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_Analytics_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------- EXPORT to Excel
@app.post("/api/export")
def export_records(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_export")
    """
    Export the given record ids to an .xlsx in the original TELCO_DATABASE layout
    (so it's interchangeable and re-importable). Only reads data; changes nothing.
    """
    import io
    import openpyxl
    import print_stage
    from datetime import datetime

    ids = data.get("ids", [])
    export_all = bool(data.get("all"))

    conn = get_conn()
    if export_all:
        ids = [r[0] for r in conn.execute(
            "SELECT id FROM licenses WHERE deleted_at IS NULL ORDER BY id")]
    if not ids:
        conn.close()
        raise HTTPException(400, "No records to export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "licenses"
    for col, h in enumerate(print_stage.STAGING_HEADERS, start=1):
        ws.cell(row=1, column=col, value=h)

    # which staging columns hold dates, so we can force a date-only display
    # (values are Python datetimes at midnight — without this, Excel shows
    # a 00:00:00 time alongside the date)
    date_cols = {i for i, f in enumerate(print_stage.STAGING_FIELDS, start=1)
                 if f in print_stage.DATE_FIELDS}

    r = 2
    for lic_id in ids:
        row = conn.execute("SELECT * FROM licenses WHERE id = ?", (lic_id,)).fetchone()
        if not row:
            continue
        rec = dict(row)
        for col, value in print_stage.staging_cells(rec):
            cell = ws.cell(row=r, column=col, value=value)
            if col in date_cols:
                cell.number_format = "yyyy-mm-dd"
        r += 1
    conn.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"RSL_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------- PRINT (single record)
@app.post("/api/print/{lic_id}")
def print_record(lic_id: int, mode: str = "preview", request: Request = None):
    """
    Preview, print, or open one record's RSL (FRONT page) via Excel.
    mode = 'preview' opens Excel's Print Preview; 'print' sends to the printer;
    'open' just shows the filled-in workbook for viewing/editing.
    Runs only on the PC where the backend runs (needs Excel + pywin32).
    """
    require_permission(request, "can_print")
    if mode not in ("preview", "print", "open"):
        raise HTTPException(400, "mode must be 'preview', 'print', or 'open'")
    try:
        import print_engine
    except ImportError:
        raise HTTPException(500, "print_engine.py not found next to main.py")
    try:
        ok = print_engine.print_one(lic_id, DB, mode)
    except FileNotFoundError as e:
        raise HTTPException(500, str(e))
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    if not ok:
        raise HTTPException(404, f"Record {lic_id} not found")
    return {"ok": True, "id": lic_id, "mode": mode}


@app.post("/api/print-word/{lic_id}")
def print_record_word(lic_id: int, mode: str = "preview", request: Request = None):
    """
    Preview, print, open, or quietly fill one record's RSL using the Word template.
    mode = 'preview' opens Word's Print Preview; 'print' sends to the printer;
    'open' just shows the filled-in document for viewing/editing; 'fill' quietly
    updates the document's fields with no visible window change — used to keep
    Word in sync while clicking through pinned stations on the map, without
    popping anything open each time.
    Runs only on the PC where the backend runs (needs Word + pywin32).
    """
    require_permission(request, "can_print")
    if mode not in ("preview", "print", "fill", "open"):
        raise HTTPException(400, "mode must be 'preview', 'print', 'open', or 'fill'")
    try:
        import print_engine_word
    except ImportError:
        raise HTTPException(500, "print_engine_word.py not found next to main.py")
    try:
        ok = print_engine_word.print_one(lic_id, DB, mode)
    except FileNotFoundError as e:
        raise HTTPException(500, str(e))
    except RuntimeError as e:
        raise HTTPException(500, str(e))
    if not ok:
        raise HTTPException(404, f"Record {lic_id} not found")
    return {"ok": True, "id": lic_id, "mode": mode}


# ---------------------------------------------------------------- IMPORT from Excel (add as new)
IMPORT_CACHE = {}   # token -> list of row dicts (parsed, awaiting confirm)

def _clean_region(v):
    return "Region II" if v not in (None, "") else v

_CLASS_MAP = {"FB (BWA)": "FB-BWA", "FB(BWA)": "FB-BWA", "FB (WDN)": "FB-WDN"}
def _clean_class(v):
    if v in (None, ""): return v
    return _CLASS_MAP.get(str(v).strip(), str(v).strip())

@app.post("/api/import")
def import_preview(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_import")
    """Parse an uploaded Excel (base64) in the export layout. Caches rows and
    returns how many valid records were found — does NOT insert yet."""
    import base64, io, datetime, secrets
    import openpyxl, print_stage

    b64 = data.get("b64", "")
    if not b64:
        raise HTTPException(400, "No file received")
    try:
        raw = base64.b64decode(b64.split(",")[-1])
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Could not read the Excel file: {e}")

    ws = wb["licenses"] if "licenses" in wb.sheetnames else wb.active

    # header (row 1) -> db field, matched by name against the export layout
    header_to_field = {}
    for h, f in zip(print_stage.STAGING_HEADERS, print_stage.STAGING_FIELDS):
        if f is not None:
            header_to_field[str(h).strip().lower()] = f
    col_field = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None: continue
        f = header_to_field.get(str(h).strip().lower())
        if f: col_field[c] = f
    if not col_field:
        raise HTTPException(400, "This file's columns don't match the export layout. "
                                 "Please import a file created by Export.")

    date_fields = print_stage.DATE_FIELDS
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for c, f in col_field.items():
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (datetime.datetime, datetime.date)):
                v = v.strftime("%Y-%m-%d")
            rec[f] = v
        # skip blank rows (no license_no and no licensee)
        if not (rec.get("license_no") or rec.get("licensee")):
            continue
        # cleanup, keeping raw
        if "region" in rec:
            rec["region_raw"] = rec.get("region"); rec["region"] = _clean_region(rec.get("region"))
        if "class_of_station" in rec:
            rec["class_of_station_raw"] = rec.get("class_of_station")
            rec["class_of_station"] = _clean_class(rec.get("class_of_station"))
        # the search list's "Type" badge reads license_status — keep it in sync
        # with the Status column so imported records show their type correctly
        if rec.get("status") and not rec.get("license_status"):
            rec["license_status"] = rec["status"]
        rows.append(rec)

    token = secrets.token_urlsafe(12)
    IMPORT_CACHE[token] = rows
    return {"ok": True, "token": token, "rows": len(rows)}

@app.post("/api/import/commit")
def import_commit(data: dict = Body(...), request: Request = None):
    require_permission(request, "can_import")
    """Insert the previously-parsed rows as NEW records (after a backup)."""
    import shutil, os, datetime
    token = data.get("token", "")
    rows = IMPORT_CACHE.get(token)
    if rows is None:
        raise HTTPException(400, "Import session expired — please pick the file again")

    # backup the database first
    backup = None
    try:
        if os.path.exists(DB):
            backup = f"telco_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy(DB, backup)
    except Exception:
        backup = None

    conn = get_conn()
    writable = set(cols(conn, "licenses")) - PROTECTED
    added = 0
    try:
        for rec in rows:
            payload = {k: v for k, v in rec.items() if k in writable and v is not None}
            if not payload:
                continue
            fields = list(payload.keys())
            sql = (f"INSERT INTO licenses ({','.join(fields)}) "
                   f"VALUES ({','.join('?'*len(fields))})")
            conn.execute(sql, [payload[f] for f in fields])
            added += 1
        conn.commit()
    finally:
        conn.close()
    IMPORT_CACHE.pop(token, None)
    log_activity(request, "import", detail=f"Imported {added} new record(s) from Excel" + (f" (backup: {backup})" if backup else ""))
    return {"ok": True, "added": added, "backup": backup}


@app.get("/api/licenses/{lic_id}/payments")
def list_payments(lic_id: int):
    conn = get_conn()
    pays = conn.execute(
        "SELECT * FROM payments WHERE license_id = ? ORDER BY year", (lic_id,)).fetchall()
    conn.close()
    return [dict(p) for p in pays]

@app.post("/api/licenses/{lic_id}/payments")
def add_payment(lic_id: int, data: dict = Body(...)):
    conn = get_conn()
    if not conn.execute("SELECT 1 FROM licenses WHERE id = ?", (lic_id,)).fetchone():
        conn.close(); raise HTTPException(404, "License not found")
    payload = clean_payload(conn, "payments", data)
    payload["license_id"] = lic_id
    fields = list(payload.keys())
    sql = (f"INSERT INTO payments ({','.join(fields)}) "
           f"VALUES ({','.join('?' * len(fields))})")
    cur = conn.execute(sql, [payload[f] for f in fields])
    conn.commit(); pid = cur.lastrowid; conn.close()
    return {"ok": True, "payment_id": pid}

@app.put("/api/payments/{pid}")
def update_payment(pid: int, data: dict = Body(...)):
    conn = get_conn()
    if not conn.execute("SELECT 1 FROM payments WHERE id = ?", (pid,)).fetchone():
        conn.close(); raise HTTPException(404, "Payment not found")
    payload = clean_payload(conn, "payments", data)
    payload.pop("license_id", None)  # don't allow re-parenting
    if not payload:
        conn.close(); raise HTTPException(400, "No valid fields supplied")
    sets = ", ".join(f"{f} = ?" for f in payload)
    conn.execute(f"UPDATE payments SET {sets} WHERE id = ?",
                 list(payload.values()) + [pid])
    conn.commit(); conn.close()
    return {"ok": True, "payment_id": pid}

@app.delete("/api/payments/{pid}")
def delete_payment(pid: int):
    conn = get_conn()
    cur = conn.execute("DELETE FROM payments WHERE id = ?", (pid,))
    conn.commit(); conn.close()
    if cur.rowcount == 0:
        raise HTTPException(404, "Payment not found")
    return {"ok": True, "deleted": pid}
